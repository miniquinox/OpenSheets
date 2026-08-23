#!/usr/bin/env python3
"""Generate the OpenSheets golden fixture corpus (Fixtures/**).

Run:  Scripts/.venv/bin/python Scripts/gen-fixtures.py --all
      Scripts/.venv/bin/python Scripts/gen-fixtures.py basic formulas
      Scripts/.venv/bin/python Scripts/gen-fixtures.py --all --with-huge

Design notes worth knowing before you edit this:

* Fixtures in `formulas/` are written with **no cached values** and then loaded
  and re-saved by headless LibreOffice. LibreOffice has to evaluate them to
  render the sheet, so the `<v>` next to every `<f>` is computed by a real
  spreadsheet engine, not by the person writing the sidecar. `validate-fixtures.py`
  then compares the sidecar against those cached values. That is what stops the
  corpus from being a circular restatement of our own assumptions.
* Fixtures in `formats/`, `structure/` and `passthrough/` are NOT round-tripped
  through LibreOffice: it rewrites custom number formats, drops veryHidden state,
  and destroys exactly the chart/pivot/VBA parts that `passthrough/` exists to
  protect.
* `hostile/` files are assembled with byte-level zip surgery. They are malformed
  on purpose, they are never executed, and they exist only so the parser can be
  proved to reject them.
"""
from __future__ import annotations

import argparse
import io
import os
import struct
import sys
import zipfile
import zlib

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from fixtures_lib import (  # noqa: E402
    FIXTURES, attach_passthrough_hashes, base_parts, cell, content_types,
    emit_csv_sidecar, emit_sidecar, fake_vba_project, fx, openpyxl_save,
    raw_zip, recalc, root_rels, shared_strings, sheet, soffice_version,
    tiny_png, workbook_rels, workbook_xml, worksheet_xml, write_bytes,
    write_json, zip_parts, MINIMAL_STYLES, NS_MAIN, NS_R, RT, XMLDECL,
)

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

LO = None  # filled in by main(), recorded in every recalculated sidecar
HAND = "hand-authored raw OOXML (the bytes ARE the expectation)"
OPX = "openpyxl authoring (literal values, no evaluation involved)"


def engine():
    return f"{LO} recalculation of formulas written without cached values"


# ==========================================================================
# basic/
# ==========================================================================

def g_basic():
    # 1. minimal
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sheet1"; ws["A1"] = 42
    openpyxl_save(wb, "basic/minimal.xlsx")
    emit_sidecar("basic/minimal.xlsx",
                 "The smallest legal workbook parses: one sheet, one numeric cell.",
                 [sheet("Sheet1", {"A1": cell("number", 42.0)},
                        dimension="A1:A1", used_range="A1:A1")],
                 verified_by=OPX)

    # 2. multi-sheet
    wb = openpyxl.Workbook()
    a = wb.active; a.title = "Alpha"; a["A1"] = "first"; a["B2"] = 1.5
    b = wb.create_sheet("Beta"); b["A1"] = "second"; b["A2"] = True
    c = wb.create_sheet("Gamma"); c["C3"] = 7
    openpyxl_save(wb, "basic/multi-sheet.xlsx")
    emit_sidecar("basic/multi-sheet.xlsx",
                 "Sheet order, sheet naming and per-sheet used ranges are independent.",
                 [sheet("Alpha", {"A1": cell("text", "first"), "B2": cell("number", 1.5)},
                        index=0, dimension="A1:B2", used_range="A1:B2"),
                  sheet("Beta", {"A1": cell("text", "second"), "A2": cell("boolean", True)},
                        index=1, dimension="A1:A2", used_range="A1:A2"),
                  sheet("Gamma", {"C3": cell("number", 7.0)},
                        index=2, dimension="C3:C3", used_range="C3:C3")],
                 verified_by=OPX)

    # 3. empty workbook
    wb = openpyxl.Workbook(); wb.active.title = "Empty"
    openpyxl_save(wb, "basic/empty-workbook.xlsx")
    emit_sidecar("basic/empty-workbook.xlsx",
                 "A sheet with zero cells yields usedRange == nil, not a crash or a 1x1 range.",
                 [sheet("Empty", {}, dimension="A1:A1", used_range=None)],
                 verified_by=OPX)

    # 4. formatting only - styled cells that hold no value at all
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Styled"
    for ref in ("A1", "B1", "C1", "A2"):
        ws[ref].fill = PatternFill("solid", fgColor="FFFF00")
        ws[ref].font = Font(bold=True, color="FF0000")
        ws[ref].alignment = Alignment(horizontal="center")
    ws["B2"].number_format = "0.000"
    openpyxl_save(wb, "basic/formatting-only.xlsx")
    emit_sidecar("basic/formatting-only.xlsx",
                 "Cells carrying only a style (no <v>, no <f>) are .empty but still occupy the used range.",
                 [sheet("Styled",
                        {"A1": cell("empty", None), "B1": cell("empty", None),
                         "C1": cell("empty", None), "A2": cell("empty", None),
                         "B2": cell("empty", None, fmt="0.000")},
                        dimension="A1:C2", used_range="A1:C2")],
                 verified_by=OPX,
                 notes="usedRange here is driven by styled-but-valueless cells; a reader that "
                       "skips cells without <v> will report the wrong dimension.")

    # 5. every CellValue case, including both string storages
    body = (
        '<row r="1">'
        '<c r="A1"><v>3.25</v></c>'
        '<c r="B1" t="s"><v>0</v></c>'
        '<c r="C1" t="inlineStr"><is><t>inline text</t></is></c>'
        '<c r="D1" t="b"><v>1</v></c>'
        '<c r="E1" t="b"><v>0</v></c>'
        '<c r="F1" t="e"><v>#N/A</v></c>'
        '<c r="G1" s="0"/>'
        '<c r="H1" t="s"><v>1</v></c>'
        "</row>"
        '<row r="2"><c r="A2"><v>-0.0000001</v></c>'
        '<c r="B2"><v>1.7976931348623157E+308</v></c>'
        '<c r="C2"><v>0</v></c></row>'
    )
    parts = base_parts([worksheet_xml(body, dimension="A1:H2")],
                       [("Types", "visible")],
                       sst=["shared one", "shared two"])
    zip_parts("basic/types.xlsx", parts)
    emit_sidecar("basic/types.xlsx",
                 "Every CellValue case and every cell type code `t` (absent/s/inlineStr/b/e) round-trips.",
                 [sheet("Types",
                        {"A1": cell("number", 3.25),
                         "B1": cell("text", "shared one"),
                         "C1": cell("text", "inline text"),
                         "D1": cell("boolean", True),
                         "E1": cell("boolean", False),
                         "F1": cell("error", "#N/A"),
                         "G1": cell("empty", None),
                         "H1": cell("text", "shared two"),
                         "A2": cell("number", -1e-07),
                         "B2": cell("number", 1.7976931348623157e308),
                         "C2": cell("number", 0.0)},
                        dimension="A1:H2", used_range="A1:H2")],
                 verified_by=HAND,
                 notes="B2 is Double.greatestFiniteMagnitude written in OOXML scientific "
                       "notation; a naive Double(String) that rejects 'E+308' will lose it.")


# ==========================================================================
# formulas/
# ==========================================================================

# label, formula (as stored, incl. the _xlfn. prefixes OOXML really uses),
# expected CellValue kind, expected value
FUNCTION_TABLE = [
    # --- math / aggregation -------------------------------------------------
    ("SUM",            "=SUM(Data!A1:A6)",                          "number", 210.0),
    ("AVERAGE",        "=AVERAGE(Data!A1:A6)",                      "number", 35.0),
    ("COUNT",          "=COUNT(Data!A1:A6)",                        "number", 6.0),
    ("COUNTA",         "=COUNTA(Data!B1:B6)",                       "number", 6.0),
    ("MIN",            "=MIN(Data!A1:A6)",                          "number", 10.0),
    ("MAX",            "=MAX(Data!A1:A6)",                          "number", 60.0),
    ("ROUND",          "=ROUND(2.567,2)",                           "number", 2.57),
    ("ROUNDUP",        "=ROUNDUP(2.111,2)",                         "number", 2.12),
    ("ROUNDDOWN",      "=ROUNDDOWN(2.999,2)",                       "number", 2.99),
    ("ABS",            "=ABS(-7.5)",                                "number", 7.5),
    ("SQRT",           "=SQRT(144)",                                "number", 12.0),
    ("POWER",          "=POWER(2,10)",                              "number", 1024.0),
    ("MOD",            "=MOD(17,5)",                                "number", 2.0),
    ("SUMPRODUCT",     "=SUMPRODUCT(Data!A1:A3,Data!C1:C3)",        "number", 170.0),
    ("SUBTOTAL",       "=SUBTOTAL(9,Data!A1:A6)",                   "number", 210.0),
    # --- logical ------------------------------------------------------------
    ("IF",             '=IF(Data!A1>5,"yes","no")',                 "text",   "yes"),
    ("IFS",            '=_xlfn.IFS(Data!A1>100,"big",Data!A1>5,"mid",TRUE(),"small")',
                                                                    "text",   "mid"),
    ("AND",            "=AND(TRUE(),Data!A1=10)",                   "boolean", True),
    ("OR",             "=OR(FALSE(),Data!A1=99)",                   "boolean", False),
    ("NOT",            "=NOT(FALSE())",                             "boolean", True),
    ("IFERROR",        '=IFERROR(1/0,"safe")',                      "text",   "safe"),
    ("SWITCH",         '=_xlfn.SWITCH(2,1,"one",2,"two","other")',  "text",   "two"),
    # --- lookup -------------------------------------------------------------
    ("VLOOKUP",        '=VLOOKUP("beta",Data!B1:C6,2,FALSE())',     "number", 2.5),
    ("HLOOKUP",        '=HLOOKUP("y",Data!A10:C11,2,FALSE())',      "number", 2.0),
    ("XLOOKUP",        '=_xlfn.XLOOKUP("gamma",Data!B1:B6,Data!C1:C6)', "number", 3.5),
    ("INDEX",          "=INDEX(Data!A1:A6,3)",                      "number", 30.0),
    ("MATCH",          "=MATCH(40,Data!A1:A6,0)",                   "number", 4.0),
    ("OFFSET",         "=OFFSET(Data!A1,2,0)",                      "number", 30.0),
    ("INDIRECT",       '=INDIRECT("Data!A2")',                      "number", 20.0),
    ("CHOOSE",         '=CHOOSE(3,"a","b","c")',                    "text",   "c"),
    # --- text ---------------------------------------------------------------
    ("CONCAT",         '=_xlfn.CONCAT(Data!B1,"-",Data!B2)',        "text",   "alpha-beta"),
    ("TEXTJOIN",       '=_xlfn.TEXTJOIN(",",TRUE(),Data!B1:B3)',    "text",   "alpha,beta,gamma"),
    ("LEFT",           "=LEFT(Data!D2,5)",                          "text",   "hello"),
    ("RIGHT",          "=RIGHT(Data!D2,5)",                         "text",   "world"),
    ("MID",            "=MID(Data!D2,7,5)",                         "text",   "world"),
    ("LEN",            "=LEN(Data!D2)",                             "number", 11.0),
    ("TRIM",           "=TRIM(Data!D1)",                            "text",   "padded"),
    ("UPPER",          "=UPPER(Data!D3)",                           "text",   "MIXED CASE"),
    ("LOWER",          "=LOWER(Data!D3)",                           "text",   "mixed case"),
    ("PROPER",         "=PROPER(Data!D3)",                          "text",   "Mixed Case"),
    ("SUBSTITUTE",     '=SUBSTITUTE(Data!D2,"world","there")',      "text",   "hello there"),
    ("REPLACE",        '=REPLACE(Data!D2,1,5,"HELLO")',             "text",   "HELLO world"),
    ("TEXT",           '=TEXT(1234.5,"#,##0.00")',                  "text",   "1,234.50"),
    ("VALUE",          '=VALUE("123.45")',                          "number", 123.45),
    # --- date (deterministic only; TODAY/NOW live in volatile.xlsx) ----------
    ("DATE",           "=DATE(2024,3,15)",                          "number", 45366.0),
    ("YEAR",           "=YEAR(DATE(2024,3,15))",                    "number", 2024.0),
    ("MONTH",          "=MONTH(DATE(2024,3,15))",                   "number", 3.0),
    ("DAY",            "=DAY(DATE(2024,3,15))",                     "number", 15.0),
    ("EDATE",          "=EDATE(DATE(2024,1,31),1)",                 "number", 45351.0),
    ("EOMONTH",        "=EOMONTH(DATE(2024,2,10),0)",               "number", 45351.0),
    ("DATEDIF",        '=DATEDIF(DATE(2024,1,1),DATE(2024,3,15),"d")', "number", 74.0),
    ("WEEKDAY",        "=WEEKDAY(DATE(2024,3,15))",                 "number", 6.0),
    ("NETWORKDAYS",    "=NETWORKDAYS(DATE(2024,3,1),DATE(2024,3,15))", "number", 11.0),
    # --- statistics ---------------------------------------------------------
    ("MEDIAN",         "=MEDIAN(Data!A1:A6)",                       "number", 35.0),
    ("STDEV.P",        "=_xlfn.STDEV.P(Data!A1:A6)",                "number", 17.0782512765993),
    ("STDEV.S",        "=_xlfn.STDEV.S(Data!A1:A6)",                "number", 18.7082869338697),
    ("VAR",            "=VAR(Data!A1:A6)",                          "number", 350.0),
    ("PERCENTILE",     "=PERCENTILE(Data!A1:A6,0.5)",               "number", 35.0),
    ("QUARTILE",       "=QUARTILE(Data!A1:A6,1)",                   "number", 22.5),
    ("RANK",           "=RANK(30,Data!A1:A6)",                      "number", 4.0),
    ("CORREL",         "=CORREL(Data!A1:A6,Data!C1:C6)",            "number", 1.0),
    # --- conditional aggregates --------------------------------------------
    ("SUMIF",          '=SUMIF(Data!B1:B6,"alpha",Data!A1:A6)',     "number", 50.0),
    ("SUMIFS",         '=SUMIFS(Data!A1:A6,Data!B1:B6,"beta")',     "number", 70.0),
    ("COUNTIF",        '=COUNTIF(Data!B1:B6,"gamma")',              "number", 2.0),
    ("COUNTIFS",       '=COUNTIFS(Data!B1:B6,"alpha",Data!A1:A6,">15")', "number", 1.0),
    ("AVERAGEIF",      '=AVERAGEIF(Data!B1:B6,"alpha",Data!A1:A6)', "number", 25.0),
    ("AVERAGEIFS",     '=AVERAGEIFS(Data!A1:A6,Data!B1:B6,"gamma")', "number", 45.0),
    ("MAXIFS",         '=_xlfn.MAXIFS(Data!A1:A6,Data!B1:B6,"beta")', "number", 50.0),
    ("MINIFS",         '=_xlfn.MINIFS(Data!A1:A6,Data!B1:B6,"beta")', "number", 20.0),
]


def _seed_data(ws):
    for i, v in enumerate([10, 20, 30, 40, 50, 60], start=1):
        ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(["alpha", "beta", "gamma", "alpha", "beta", "gamma"], start=1):
        ws.cell(row=i, column=2, value=v)
    for i, v in enumerate([1.5, 2.5, 3.5, 4.5, 5.5, 6.5], start=1):
        ws.cell(row=i, column=3, value=v)
    ws["D1"] = "  padded  "
    ws["D2"] = "hello world"
    ws["D3"] = "mIxEd CaSe"
    for i, v in enumerate(["x", "y", "z"], start=1):
        ws.cell(row=10, column=i, value=v)
    for i, v in enumerate([1, 2, 3], start=1):
        ws.cell(row=11, column=i, value=v)


def g_formulas():
    # --- 1. the function table -------------------------------------------
    wb = openpyxl.Workbook()
    data = wb.active; data.title = "Data"; _seed_data(data)
    calc = wb.create_sheet("Calc")
    cells = {}
    for i, (label, formula, kind, expected) in enumerate(FUNCTION_TABLE, start=1):
        calc.cell(row=i, column=1, value=label)
        calc.cell(row=i, column=2, value=formula)
        cells[f"A{i}"] = cell("text", label)
        cells[f"B{i}"] = cell(kind, expected, formula=formula.lstrip("="), fmt=None)
    openpyxl_save(wb, "formulas/functions.xlsx")
    recalc(fx("formulas/functions.xlsx"))
    emit_sidecar("formulas/functions.xlsx",
                 "Every function family in PLAN.md 5.3 evaluated once, with the result a real "
                 "engine cached next to the formula. Column A names the function, column B is it.",
                 [sheet("Data", {"A1": cell("number", 10.0), "B1": cell("text", "alpha"),
                                 "C6": cell("number", 6.5), "D2": cell("text", "hello world"),
                                 "A11": cell("number", 1.0)},
                        index=0, used_range="A1:D11"),
                  sheet("Calc", cells, index=1,
                        used_range=f"A1:B{len(FUNCTION_TABLE)}")],
                 verified_by=engine(),
                 notes="Formula TEXT is stored with the OOXML `_xlfn.` prefix for every "
                       "post-2007 function (IFS, SWITCH, CONCAT, TEXTJOIN, XLOOKUP, MAXIFS, "
                       "MINIFS, STDEV.P, STDEV.S). A3 must strip/map that prefix; the user "
                       "never sees it in Excel's formula bar.")

    # --- 2. volatile functions (value intentionally unasserted) -----------
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Volatile"
    ws["A1"] = "TODAY"; ws["B1"] = "=TODAY()"
    ws["A2"] = "NOW";   ws["B2"] = "=NOW()"
    ws["A3"] = "RAND-free check"; ws["B3"] = "=YEAR(TODAY())>=2024"
    openpyxl_save(wb, "formulas/volatile.xlsx")
    recalc(fx("formulas/volatile.xlsx"))
    emit_sidecar("formulas/volatile.xlsx",
                 "TODAY()/NOW() parse and carry a cached value that must NOT be asserted "
                 "against a constant - the corpus cannot pin a clock.",
                 [sheet("Volatile",
                        {"A1": cell("text", "TODAY"),
                         "B1": cell("number", None, formula="TODAY()", fmt=None),
                         "A2": cell("text", "NOW"),
                         "B2": cell("number", None, formula="NOW()", fmt=None),
                         "B3": cell("boolean", True, formula="YEAR(TODAY())>=2024", fmt=None)},
                        used_range="A1:B3")],
                 verified_by=engine(),
                 skip_checks=["cellValue:Volatile!B1", "cellValue:Volatile!B2"],
                 notes="value:null means 'a value must exist, its content is not asserted'.")

    # --- 3. cross-sheet references ---------------------------------------
    wb = openpyxl.Workbook()
    a = wb.active; a.title = "Source"; a["A1"] = 100; a["A2"] = 250
    b = wb.create_sheet("Middle"); b["A1"] = "=Source!A1*2"; b["A2"] = "=SUM(Source!A1:A2)"
    c = wb.create_sheet("Far Away")  # space in the name -> quoted reference
    c["A1"] = 7
    b["A3"] = "='Far Away'!A1+1"
    openpyxl_save(wb, "formulas/cross-sheet.xlsx")
    recalc(fx("formulas/cross-sheet.xlsx"))
    emit_sidecar("formulas/cross-sheet.xlsx",
                 "References across sheets, including a sheet whose name needs single-quoting.",
                 [sheet("Source", {"A1": cell("number", 100.0), "A2": cell("number", 250.0)},
                        index=0, used_range="A1:A2"),
                  sheet("Middle",
                        {"A1": cell("number", 200.0, formula="Source!A1*2", fmt=None),
                         "A2": cell("number", 350.0, formula="SUM(Source!A1:A2)", fmt=None),
                         "A3": cell("number", 8.0, formula="'Far Away'!A1+1", fmt=None)},
                        index=1, used_range="A1:A3"),
                  sheet("Far Away", {"A1": cell("number", 7.0)}, index=2, used_range="A1:A1")],
                 verified_by=engine())

    # --- 4. defined names --------------------------------------------------
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Budget"
    for i, v in enumerate([100, 200, 300], start=1):
        ws.cell(row=i, column=1, value=v)
    ws["C1"] = 0.08
    dn = openpyxl.workbook.defined_name.DefinedName("Revenue", attr_text="Budget!$A$1:$A$3")
    wb.defined_names["Revenue"] = dn
    wb.defined_names["GrowthRate"] = openpyxl.workbook.defined_name.DefinedName(
        "GrowthRate", attr_text="Budget!$C$1")
    ws["E1"] = "=SUM(Revenue)"
    ws["E2"] = "=SUM(Revenue)*(1+GrowthRate)"
    openpyxl_save(wb, "formulas/defined-names.xlsx")
    recalc(fx("formulas/defined-names.xlsx"))
    emit_sidecar("formulas/defined-names.xlsx",
                 "Workbook-scoped defined names resolve inside formulas and survive as ranges.",
                 [sheet("Budget",
                        {"A1": cell("number", 100.0), "C1": cell("number", 0.08),
                         "E1": cell("number", 600.0, formula="SUM(Revenue)", fmt=None),
                         "E2": cell("number", 648.0, formula="SUM(Revenue)*(1+GrowthRate)", fmt=None)},
                        used_range="A1:E3")],
                 defined_names={"Revenue": "Budget!$A$1:$A$3", "GrowthRate": "Budget!$C$1"},
                 verified_by=engine())

    # --- 5. formulas that evaluate to each error kind ---------------------
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Errors"
    # NOTE the labels deliberately omit the leading '#': writing the literal
    # text "#DIV/0!" into a cell makes it an ERROR cell, not a text cell.
    err_rows = [
        ("DIV/0", "#DIV/0!", "=1/0"),
        ("N/A",   "#N/A",    "=NA()"),
        ("VALUE", "#VALUE!", '=SQRT("not a number")'),
        ("NAME",  "#NAME?",  "=nosuchfunction()"),
        ("NUM",   "#NUM!",   "=ASIN(5)"),
        ("NULL",  "#NULL!",  "=SUM(A1:A2 C1:C2)"),
        ("REF",   "#REF!",   "=INDEX(A1:A3,99)"),
    ]
    cells = {}
    for i, (label, err, f) in enumerate(err_rows, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=f)
        cells[f"A{i}"] = cell("text", label)
        cells[f"B{i}"] = cell("error", err, formula=f.lstrip("="), fmt=None)
    openpyxl_save(wb, "formulas/error-formulas.xlsx")
    recalc(fx("formulas/error-formulas.xlsx"))
    emit_sidecar("formulas/error-formulas.xlsx",
                 "All seven Excel error kinds arise from real evaluation and are cached as t=\"e\".",
                 [sheet("Errors", cells, used_range="A1:B7")],
                 verified_by=engine(),
                 notes="Engines DISAGREE on which error a bad argument produces: SQRT(-1) is "
                       "#NUM! in Excel but #VALUE! in LibreOffice, and OFFSET past the edge is "
                       "#REF! in Excel but #VALUE! in LibreOffice. The formulas here were chosen "
                       "because both engines agree on them. A3 must NOT treat a recalculation "
                       "oracle as authoritative for error KIND - only for value.")

    # --- 6. cached errors with NO formula (hand-authored bytes) -----------
    body = "".join(
        f'<row r="{i}"><c r="A{i}" t="s"><v>{i-1}</v></c>'
        f'<c r="B{i}" t="e"><v>{e}</v></c></row>'
        for i, e in enumerate(["#DIV/0!", "#REF!", "#NAME?", "#VALUE!",
                               "#N/A", "#NULL!", "#NUM!"], start=1))
    parts = base_parts([worksheet_xml(body, dimension="A1:B7")], [("Cached", "visible")],
                       sst=["#DIV/0!", "#REF!", "#NAME?", "#VALUE!", "#N/A", "#NULL!", "#NUM!"])
    zip_parts("formulas/cached-errors.xlsx", parts)
    emit_sidecar("formulas/cached-errors.xlsx",
                 "A literal cached error value with no <f> at all - Excel writes these after a "
                 "delete, and a reader that only creates errors from formulas will drop them.",
                 [sheet("Cached",
                        {f"B{i}": cell("error", e) for i, e in enumerate(
                            ["#DIV/0!", "#REF!", "#NAME?", "#VALUE!", "#N/A", "#NULL!", "#NUM!"],
                            start=1)},
                        dimension="A1:B7", used_range="A1:B7")],
                 verified_by=HAND)

    # --- 7. shared formulas -----------------------------------------------
    rows = []
    for r in range(1, 9):
        f = ('<f t="shared" ref="B1:B8" si="0">A1*2</f>' if r == 1
             else '<f t="shared" si="0"/>')
        rows.append(f'<row r="{r}"><c r="A{r}"><v>{r}</v></c>'
                    f'<c r="B{r}">{f}<v>{r*2}</v></c></row>')
    parts = base_parts([worksheet_xml("".join(rows), dimension="A1:B8")], [("Shared", "visible")])
    zip_parts("formulas/shared-formulas.xlsx", parts)
    emit_sidecar("formulas/shared-formulas.xlsx",
                 "A <f t=\"shared\"> master plus 7 empty followers must expand to per-cell "
                 "formula text with the refs translated (B3 is A3*2, not A1*2).",
                 [sheet("Shared",
                        {**{f"A{r}": cell("number", float(r)) for r in range(1, 9)},
                         **{f"B{r}": cell("number", float(r * 2), formula=f"A{r}*2")
                            for r in range(1, 9)}},
                        dimension="A1:B8", used_range="A1:B8")],
                 verified_by=HAND,
                 skip_checks=["formulaText:Shared"],
                 notes="skipChecks formulaText: the raw XML holds `<f t=\"shared\" si=\"0\"/>` "
                       "for B2..B8; the sidecar states the EXPANDED text A1's reader must produce.")

    # --- 8. array formulas -------------------------------------------------
    body = (
        '<row r="1"><c r="A1"><v>1</v></c><c r="B1"><v>10</v></c>'
        '<c r="D1"><f t="array" ref="D1:D3">A1:A3*B1:B3</f><v>10</v></c>'
        '<c r="F1"><f t="array" ref="F1">SUM(A1:A3*B1:B3)</f><v>140</v></c></row>'
        '<row r="2"><c r="A2"><v>2</v></c><c r="B2"><v>20</v></c>'
        '<c r="D2"><v>40</v></c></row>'
        '<row r="3"><c r="A3"><v>3</v></c><c r="B3"><v>30</v></c>'
        '<c r="D3"><v>90</v></c></row>'
    )
    parts = base_parts([worksheet_xml(body, dimension="A1:F3")], [("Arrays", "visible")])
    zip_parts("formulas/array-formulas.xlsx", parts)
    emit_sidecar("formulas/array-formulas.xlsx",
                 "CSE array formulas: the <f t=\"array\" ref=\"D1:D3\"> master owns a 3-cell "
                 "spill whose other cells carry only <v>, plus a single-cell array in F1.",
                 [sheet("Arrays",
                        {"D1": cell("number", 10.0, formula="A1:A3*B1:B3"),
                         "D2": cell("number", 40.0),
                         "D3": cell("number", 90.0),
                         "F1": cell("number", 140.0, formula="SUM(A1:A3*B1:B3)")},
                        dimension="A1:F3", used_range="A1:F3")],
                 verified_by=HAND,
                 notes="D2/D3 legitimately have a value and no formula. Writing them back as "
                       "plain constants would silently break the array on the next Excel open.")

    # --- 9. circular reference ---------------------------------------------
    body = ('<row r="1"><c r="A1"><f>B1+1</f><v>0</v></c>'
            '<c r="B1"><f>A1+1</f><v>0</v></c>'
            '<c r="C1"><f>C1</f><v>0</v></c></row>')
    parts = base_parts([worksheet_xml(body, dimension="A1:C1")], [("Loop", "visible")])
    parts["xl/workbook.xml"] = workbook_xml(
        [("Loop", "visible")], extra_head='<calcPr calcId="0" iterate="1" iterateCount="100"/>')
    zip_parts("formulas/circular-reference.xlsx", parts)
    emit_sidecar("formulas/circular-reference.xlsx",
                 "A 2-cell cycle and a 1-cell self-reference must yield #CIRCULAR, never a hang.",
                 [sheet("Loop",
                        {"A1": cell("number", 0.0, formula="B1+1"),
                         "B1": cell("number", 0.0, formula="A1+1"),
                         "C1": cell("number", 0.0, formula="C1")},
                        dimension="A1:C1", used_range="A1:C1")],
                 verified_by=HAND,
                 notes="Reading is unaffected (cached 0 is displayed). It is A3's recalc that "
                       "must detect the cycle. calcPr iterate=1 is present on purpose.")

    # --- 10. external workbook link ---------------------------------------
    ext_link = (XMLDECL + f'<externalLink xmlns="{NS_MAIN}" xmlns:r="{NS_R}">'
                '<externalBook r:id="rId1"><sheetNames><sheetName val="Sheet1"/></sheetNames>'
                '<sheetDataSet><sheetData sheetId="0">'
                '<row r="1"><cell r="A1"><v>999</v></cell></row>'
                "</sheetData></sheetDataSet></externalBook></externalLink>")
    ext_rels = (XMLDECL + '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                f'<Relationship Id="rId1" Type="{RT}/externalLinkPath" '
                'Target="file:///Users/nobody/never-fetched.xlsx" TargetMode="External"/>'
                "</Relationships>")
    body = ('<row r="1"><c r="A1"><f>[1]Sheet1!A1</f><v>999</v></c>'
            '<c r="B1"><f>[1]Sheet1!A1*2</f><v>1998</v></c>'
            '<c r="C1"><f>SUM([1]Sheet1!A1:A5)</f><v>999</v></c></row>')
    parts = base_parts([worksheet_xml(body, dimension="A1:C1")], [("Linked", "visible")])
    parts["xl/workbook.xml"] = workbook_xml(
        [("Linked", "visible")],
        defined_names='<externalReferences><externalReference r:id="rIdExt"/></externalReferences>')
    parts["xl/_rels/workbook.xml.rels"] = workbook_rels(
        1, extra=f'<Relationship Id="rIdExt" Type="{RT}/externalLink" '
                 'Target="externalLinks/externalLink1.xml"/>')
    parts["[Content_Types].xml"] = content_types(
        1, extra='<Override PartName="/xl/externalLinks/externalLink1.xml" '
                 'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"/>')
    parts["xl/externalLinks/externalLink1.xml"] = ext_link
    parts["xl/externalLinks/_rels/externalLink1.xml.rels"] = ext_rels
    zip_parts("formulas/external-link.xlsx", parts)
    emit_sidecar("formulas/external-link.xlsx",
                 "Formulas referencing another workbook ([1]Sheet1!A1) must set the "
                 ".externalLink cell flag and must NEVER be resolved or fetched.",
                 [sheet("Linked",
                        {"A1": cell("number", 999.0, formula="[1]Sheet1!A1",
                                    flags=["externalLink"]),
                         "B1": cell("number", 1998.0, formula="[1]Sheet1!A1*2",
                                    flags=["externalLink"]),
                         "C1": cell("number", 999.0, formula="SUM([1]Sheet1!A1:A5)",
                                    flags=["externalLink"])},
                        dimension="A1:C1", used_range="A1:C1")],
                 verified_by=HAND,
                 mustNotHappen=["any filesystem or network read of the externalLinkPath target"],
                 notes="The link target is a file:/// URL that does not exist. If a test observes "
                       "an open() on it, the reader is broken (PLAN.md 7.3).")


# ==========================================================================
# formats/
# ==========================================================================

BUILTIN_NUMFMTS = {
    0: "General", 1: "0", 2: "0.00", 3: "#,##0", 4: "#,##0.00",
    9: "0%", 10: "0.00%", 11: "0.00E+00", 12: "# ?/?", 13: "# ??/??",
    14: "mm-dd-yy", 15: "d-mmm-yy", 16: "d-mmm", 17: "mmm-yy",
    18: "h:mm AM/PM", 19: "h:mm:ss AM/PM", 20: "h:mm", 21: "h:mm:ss",
    22: "m/d/yy h:mm",
    37: "#,##0 ;(#,##0)", 38: "#,##0 ;[Red](#,##0)",
    39: "#,##0.00;(#,##0.00)", 40: "#,##0.00;[Red](#,##0.00)",
    45: "mm:ss", 46: "[h]:mm:ss", 47: "mmss.0", 48: "##0.0E+0", 49: "@",
}

CUSTOM_NUMFMTS = [
    ("0.000", 1234.5678),
    ("#,##0", 1234567.0),
    ("#,##0.00", 1234.5),
    ("0%", 0.4267),
    ("0.00%", 0.4267),
    ("0.00E+00", 12345.6789),
    ("$#,##0.00", 1999.5),
    ('$#,##0.00;[Red]($#,##0.00)', -1999.5),
    ('#,##0.00;[Red]-#,##0.00', -42.5),
    ('0.00_);[Red](0.00)', -7.25),
    ("@", "0012"),
    ('"EUR "#,##0.00', 88.25),
    ("yyyy-mm-dd", 45366.0),
    ("dd/mm/yyyy hh:mm", 45366.5),
    ("[h]:mm:ss", 1.5),
    ('[>=1000]#,##0,"k";#,##0', 12500.0),
    ('_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-', 3.5),
]


def g_formats():
    # 1. all built-in number format ids that Excel never writes into the file
    body_rows, cells, xfs = [], {}, []
    for i, (nid, code) in enumerate(sorted(BUILTIN_NUMFMTS.items())):
        xfs.append(f'<xf numFmtId="{nid}" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>')
        r = i + 1
        body_rows.append(f'<row r="{r}"><c r="A{r}"><v>{nid}</v></c>'
                         f'<c r="B{r}" s="{i}"><v>45366.625</v></c></row>')
        cells[f"A{r}"] = cell("number", float(nid))
        cells[f"B{r}"] = cell("number", 45366.625, fmt=code)
    styles = MINIMAL_STYLES.replace(
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>',
        f'<cellXfs count="{len(xfs)}">' + "".join(xfs) + "</cellXfs>")
    parts = base_parts([worksheet_xml("".join(body_rows), dimension=f"A1:B{len(xfs)}")],
                       [("Builtins", "visible")], styles=styles)
    zip_parts("formats/builtin-numfmts.xlsx", parts)
    emit_sidecar("formats/builtin-numfmts.xlsx",
                 "The ~25 built-in numFmtIds that are IMPLICIT - they never appear in "
                 "<numFmts>, so a reader without the hardcoded table renders them all as General.",
                 [sheet("Builtins", cells, dimension=f"A1:B{len(xfs)}",
                        used_range=f"A1:B{len(xfs)}")],
                 verified_by=HAND,
                 notes="Column A is the numFmtId; column B is the same serial (45366.625 = "
                       "2024-03-15 15:00) rendered through it.")

    # 2. custom formats declared in <numFmts>
    numfmts, xfs, body_rows, cells = [], [], [], {}
    sst = []
    for i, (code, value) in enumerate(CUSTOM_NUMFMTS):
        nid = 164 + i
        esc = (code.replace("&", "&amp;").replace("<", "&lt;")
               .replace(">", "&gt;").replace('"', "&quot;"))
        numfmts.append(f'<numFmt numFmtId="{nid}" formatCode="{esc}"/>')
        xfs.append(f'<xf numFmtId="{nid}" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>')
        r = i + 1
        if isinstance(value, str):
            sst.append(value)
            vcell = f'<c r="B{r}" s="{i+1}" t="s"><v>{len(sst)-1}</v></c>'
            cells[f"B{r}"] = cell("text", value, fmt=code)
        else:
            vcell = f'<c r="B{r}" s="{i+1}"><v>{value!r}</v></c>'
            cells[f"B{r}"] = cell("number", float(value), fmt=code)
        body_rows.append(f'<row r="{r}"><c r="A{r}" t="inlineStr"><is><t>{esc}</t></is></c>{vcell}</row>')
        cells[f"A{r}"] = cell("text", code)
    styles = (MINIMAL_STYLES
              .replace('<numFmts count="0"/>',
                       f'<numFmts count="{len(numfmts)}">' + "".join(numfmts) + "</numFmts>")
              .replace('<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>',
                       f'<cellXfs count="{len(xfs)+1}">'
                       '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
                       + "".join(xfs) + "</cellXfs>"))
    parts = base_parts([worksheet_xml("".join(body_rows), dimension=f"A1:B{len(xfs)}")],
                       [("Custom", "visible")], sst=sst, styles=styles)
    zip_parts("formats/custom-numfmts.xlsx", parts)
    emit_sidecar("formats/custom-numfmts.xlsx",
                 "Custom format codes: negative-in-red, currency, percent, scientific, "
                 "accounting alignment, conditional sections, elapsed time, and the text "
                 "placeholder @.",
                 [sheet("Custom", cells, dimension=f"A1:B{len(xfs)}",
                        used_range=f"A1:B{len(xfs)}")],
                 verified_by=HAND,
                 notes="Column A repeats the format code as text so a failing render is "
                       "readable side by side.")

    # 3/4. the same wall-clock dates in the 1900 and 1904 epochs
    dates = ["1900-01-01", "1900-03-01", "1970-01-01", "2024-03-15", "2099-12-31"]
    serial1900 = [1.0, 61.0, 25569.0, 45366.0, 73050.0]
    serial1904 = [s - 1462.0 for s in serial1900]
    for tag, serials, flag in (("1900", serial1900, False), ("1904", serial1904, True)):
        rows, cells = [], {}
        for i, (d, s) in enumerate(zip(dates, serials), start=1):
            rows.append(f'<row r="{i}"><c r="A{i}" t="inlineStr"><is><t>{d}</t></is></c>'
                        f'<c r="B{i}" s="1"><v>{s!r}</v></c></row>')
            cells[f"A{i}"] = cell("text", d)
            cells[f"B{i}"] = cell("number", s, fmt="yyyy-mm-dd")
        styles = (MINIMAL_STYLES
                  .replace('<numFmts count="0"/>',
                           '<numFmts count="1"><numFmt numFmtId="164" formatCode="yyyy-mm-dd"/></numFmts>')
                  .replace('<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>',
                           '<cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
                           '<xf numFmtId="164" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/></cellXfs>'))
        parts = base_parts([worksheet_xml("".join(rows), dimension="A1:B5")],
                           [(f"Dates{tag}", "visible")], styles=styles, date1904=flag)
        zip_parts(f"formats/dates-{tag}.xlsx", parts)
        emit_sidecar(f"formats/dates-{tag}.xlsx",
                     f"The {tag} date system. Column A is the wall-clock date as text, column B "
                     f"the serial. dates-1900 and dates-1904 must render IDENTICAL dates.",
                     [sheet(f"Dates{tag}", cells, dimension="A1:B5", used_range="A1:B5")],
                     verified_by=HAND, date_system=int(tag),
                     twinFixture=f"formats/dates-{'1904' if tag == '1900' else '1900'}.xlsx",
                     notes="The 1904 serials are exactly the 1900 serials minus 1462.")

    # 5. the Lotus 1-2-3 leap-year bug
    rows, cells = [], {}
    lotus = [(59, "1900-02-28", "real day"),
             (60, "1900-02-29", "DOES NOT EXIST - Lotus 1-2-3 compatibility bug, Excel keeps it"),
             (61, "1900-03-01", "real day, off-by-one vs a naive epoch+n calculation")]
    for i, (s, d, why) in enumerate(lotus, start=1):
        rows.append(f'<row r="{i}"><c r="A{i}" s="1"><v>{s}</v></c>'
                    f'<c r="B{i}" t="inlineStr"><is><t>{d}</t></is></c>'
                    f'<c r="C{i}" t="inlineStr"><is><t>{why}</t></is></c></row>')
        cells[f"A{i}"] = cell("number", float(s), fmt="yyyy-mm-dd")
        cells[f"B{i}"] = cell("text", d)
        cells[f"C{i}"] = cell("text", why)
    styles = (MINIMAL_STYLES
              .replace('<numFmts count="0"/>',
                       '<numFmts count="1"><numFmt numFmtId="164" formatCode="yyyy-mm-dd"/></numFmts>')
              .replace('<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>',
                       '<cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
                       '<xf numFmtId="164" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/></cellXfs>'))
    parts = base_parts([worksheet_xml("".join(rows), dimension="A1:C3")],
                       [("LotusBug", "visible")], styles=styles)
    zip_parts("formats/serial-60-lotus-bug.xlsx", parts)
    emit_sidecar("formats/serial-60-lotus-bug.xlsx",
                 "Serial 60 is 1900-02-29, a date that never happened. Any date conversion that "
                 "does not special-case serials <= 60 is one day wrong for all of Jan/Feb 1900.",
                 [sheet("LotusBug", cells, dimension="A1:C3", used_range="A1:C3")],
                 verified_by=HAND,
                 notes="Expected rendering: A1=1900-02-28, A2=1900-02-29 (fictional), A3=1900-03-01.")

    # 6. numbers stored as text, and text that looks numeric
    sst = ["0012", "1.23", "  42  ", "1e5", "TRUE", "=1+1", "+1-800-555", "-", "00:30"]
    rows, cells = [], {}
    for i, v in enumerate(sst, start=1):
        esc = v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        rows.append(f'<row r="{i}"><c r="A{i}" s="1" t="s"><v>{i-1}</v></c></row>')
        cells[f"A{i}"] = cell("text", v, fmt="@")
    styles = MINIMAL_STYLES.replace(
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>',
        '<cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        '<xf numFmtId="49" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/></cellXfs>')
    parts = base_parts([worksheet_xml("".join(rows), dimension="A1:A9")],
                       [("AsText", "visible")], sst=sst, styles=styles)
    zip_parts("formats/text-format.xlsx", parts)
    emit_sidecar("formats/text-format.xlsx",
                 "numFmt @ forces text. '0012' must keep its leading zeros and '=1+1' must stay "
                 "a string, not become a formula, on read OR on re-save.",
                 [sheet("AsText", cells, dimension="A1:A9", used_range="A1:A9")],
                 verified_by=HAND)


# ==========================================================================
# structure/
# ==========================================================================

def _styles_with_xfs(xfs, numfmts=""):
    s = MINIMAL_STYLES
    if numfmts:
        s = s.replace('<numFmts count="0"/>',
                      f'<numFmts count="{numfmts.count("<numFmt ")}">{numfmts}</numFmts>')
    return s.replace(
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>',
        f'<cellXfs count="{len(xfs)}">' + "".join(xfs) + "</cellXfs>")


def g_structure():
    # 1. merged cells
    body = ('<row r="1"><c r="A1" t="inlineStr"><is><t>title spans A1:D1</t></is></c></row>'
            '<row r="2"><c r="A2" t="inlineStr"><is><t>block</t></is></c></row>'
            '<row r="5"><c r="A5" t="inlineStr"><is><t>tall</t></is></c>'
            '<c r="F5"><v>1</v></c></row>')
    merges = ('<mergeCells count="3"><mergeCell ref="A1:D1"/><mergeCell ref="A2:B3"/>'
              '<mergeCell ref="A5:A8"/></mergeCells>')
    parts = base_parts([worksheet_xml(body, dimension="A1:F8", merges=merges)],
                       [("Merged", "visible")])
    zip_parts("structure/merged-cells.xlsx", parts)
    emit_sidecar("structure/merged-cells.xlsx",
                 "Horizontal, rectangular and vertical merges. Only the anchor cell carries a "
                 "value; the covered cells are absent from the XML entirely.",
                 [sheet("Merged",
                        {"A1": cell("text", "title spans A1:D1"),
                         "A2": cell("text", "block"),
                         "A5": cell("text", "tall"),
                         "F5": cell("number", 1.0)},
                        dimension="A1:F8", used_range="A1:F8",
                        merges=["A1:D1", "A2:B3", "A5:A8"])],
                 verified_by=HAND)

    # 2. frozen panes
    views = ('<sheetViews><sheetView workbookViewId="0">'
             '<pane xSplit="2" ySplit="1" topLeftCell="C2" activePane="bottomRight" state="frozen"/>'
             '<selection pane="bottomRight" activeCell="C2" sqref="C2"/>'
             "</sheetView></sheetViews>")
    body = "".join(
        f'<row r="{r}">' + "".join(
            f'<c r="{get_column_letter(c)}{r}"><v>{r*100+c}</v></c>' for c in range(1, 6)
        ) + "</row>" for r in range(1, 11))
    parts = base_parts([worksheet_xml(body, dimension="A1:E10", views=views)],
                       [("Frozen", "visible")])
    zip_parts("structure/frozen-panes.xlsx", parts)
    emit_sidecar("structure/frozen-panes.xlsx",
                 "Two frozen header rows/columns (xSplit=2 ySplit=1 state=frozen).",
                 [sheet("Frozen",
                        {"A1": cell("number", 101.0), "E10": cell("number", 1005.0)},
                        dimension="A1:E10", used_range="A1:E10",
                        frozen={"columns": 2, "rows": 1, "topLeftCell": "C2", "state": "frozen"})],
                 verified_by=HAND)

    # 3. split (not frozen) panes - the same element, different `state`
    views = ('<sheetViews><sheetView workbookViewId="0">'
             '<pane xSplit="2130" ySplit="900" topLeftCell="D5" activePane="bottomRight"/>'
             "</sheetView></sheetViews>")
    parts = base_parts([worksheet_xml(body, dimension="A1:E10", views=views)],
                       [("Split", "visible")])
    zip_parts("structure/split-panes.xlsx", parts)
    emit_sidecar("structure/split-panes.xlsx",
                 "A SPLIT pane: same <pane> element but state is absent, and xSplit/ySplit are "
                 "1/20th-of-a-point offsets, not counts. Reading them as counts gives a "
                 "2130-column freeze.",
                 [sheet("Split",
                        {"A1": cell("number", 101.0), "E10": cell("number", 1005.0)},
                        dimension="A1:E10", used_range="A1:E10",
                        frozen=None,
                        split={"xSplitTwips": 2130, "ySplitTwips": 900, "topLeftCell": "D5"})],
                 verified_by=HAND)

    # 4. a merge straddling the frozen boundary
    views = ('<sheetViews><sheetView workbookViewId="0">'
             '<pane xSplit="2" ySplit="2" topLeftCell="C3" activePane="bottomRight" state="frozen"/>'
             "</sheetView></sheetViews>")
    body = ('<row r="1"><c r="A1" t="inlineStr"><is><t>straddles the freeze</t></is></c></row>'
            '<row r="3"><c r="A3" t="inlineStr"><is><t>also straddles</t></is></c></row>')
    merges = ('<mergeCells count="2"><mergeCell ref="A1:D4"/><mergeCell ref="B2:C6"/></mergeCells>')
    parts = base_parts([worksheet_xml(body, dimension="A1:D6", views=views, merges=merges)],
                       [("Straddle", "visible")])
    zip_parts("structure/merge-across-frozen-boundary.xlsx", parts)
    emit_sidecar("structure/merge-across-frozen-boundary.xlsx",
                 "A merged range that crosses the frozen split in both axes - the renderer must "
                 "clip it per pane instead of drawing it four times.",
                 [sheet("Straddle",
                        {"A1": cell("text", "straddles the freeze"),
                         "A3": cell("text", "also straddles")},
                        dimension="A1:D6", used_range="A1:D6",
                        merges=["A1:D4", "B2:C6"],
                        frozen={"columns": 2, "rows": 2, "topLeftCell": "C3", "state": "frozen"})],
                 verified_by=HAND)

    # 5. hidden and very hidden sheets
    names = [("Visible", "visible"), ("Hidden", "hidden"),
             ("VeryHidden", "veryHidden"), ("AlsoVisible", "visible")]
    bodies = [worksheet_xml(f'<row r="1"><c r="A1" t="inlineStr"><is><t>{n}</t></is></c></row>',
                            dimension="A1:A1") for n, _ in names]
    parts = base_parts(bodies, names)
    zip_parts("structure/hidden-sheets.xlsx", parts)
    emit_sidecar("structure/hidden-sheets.xlsx",
                 "visible / hidden / veryHidden. veryHidden cannot be unhidden from Excel's UI "
                 "and is the one most libraries silently drop on save.",
                 [sheet(n, {"A1": cell("text", n)}, index=i, visibility=st,
                        dimension="A1:A1", used_range="A1:A1")
                  for i, (n, st) in enumerate(names)],
                 verified_by=HAND)

    # 6. custom column widths and row heights
    cols = ('<cols>'
            '<col min="1" max="1" width="3.5" customWidth="1"/>'
            '<col min="2" max="2" width="42.75" customWidth="1"/>'
            '<col min="3" max="10" width="12" customWidth="1"/>'
            '<col min="11" max="11" width="0" hidden="1" customWidth="1"/>'
            '<col min="12" max="16384" width="8.7109375"/>'
            "</cols>")
    rows = ('<row r="1" ht="40.5" customHeight="1"><c r="A1"><v>1</v></c></row>'
            '<row r="2"><c r="A2"><v>2</v></c></row>'
            '<row r="3" ht="8" customHeight="1"><c r="A3"><v>3</v></c></row>'
            '<row r="4" hidden="1"><c r="A4"><v>4</v></c></row>')
    fmtpr = ""
    parts = base_parts([worksheet_xml(rows, dimension="A1:K4", cols=cols)],
                       [("Sizes", "visible")])
    zip_parts("structure/col-widths-row-heights.xlsx", parts)
    emit_sidecar("structure/col-widths-row-heights.xlsx",
                 "Sparse column widths (including a run covering all 16384 columns and a hidden "
                 "column) and per-row heights. RunLengthArray must not materialise 16384 doubles.",
                 [sheet("Sizes",
                        {f"A{r}": cell("number", float(r)) for r in range(1, 5)},
                        dimension="A1:K4", used_range="A1:A4",
                        col_widths={"1": 3.5, "2": 42.75, "3-10": 12.0, "11": 0.0,
                                    "12-16384": 8.7109375},
                        row_heights={"1": 40.5, "3": 8.0})],
                 verified_by=HAND,
                 hiddenColumns=[11], hiddenRows=[4],
                 notes="usedRange is A1:A4 even though <dimension> claims A1:K4 - dimension is a "
                       "capacity hint from the producer and is routinely wrong.")

    # 7. hostile-but-legal sheet names
    names = [("\U0001F4CA Data", "visible"),
             ("ورقة العمل", "visible"),
             ("ThisSheetNameIsExactly31Charsxx", "visible"),
             ("Report", "visible"),
             ("report", "visible"),
             ("A B  C", "visible")]
    bodies = [worksheet_xml('<row r="1"><c r="A1"><v>1</v></c></row>', dimension="A1:A1")
              for _ in names]
    parts = base_parts(bodies, names)
    zip_parts("structure/sheet-names-unicode.xlsx", parts)
    emit_sidecar("structure/sheet-names-unicode.xlsx",
                 "Emoji, RTL Arabic, exactly 31 characters, two names differing only by case, "
                 "and a name with a double space. Case-insensitive sheet lookup breaks here.",
                 [sheet(n, {"A1": cell("number", 1.0)}, index=i, visibility=st,
                        dimension="A1:A1", used_range="A1:A1")
                  for i, (n, st) in enumerate(names)],
                 verified_by=HAND,
                 notes="Excel forbids creating Report/report in its UI but happily OPENS a file "
                       "that contains both. Sheet lookup must be case-SENSITIVE and ordinal.")

    # 8. rows and cells emitted out of order
    body = ('<row r="7"><c r="C7"><v>73</v></c><c r="A7"><v>71</v></c></row>'
            '<row r="2"><c r="B2"><v>22</v></c></row>'
            '<row r="9"><c r="E9"><v>95</v></c><c r="B9"><v>92</v></c><c r="D9"><v>94</v></c></row>'
            '<row r="1"><c r="A1"><v>11</v></c></row>')
    parts = base_parts([worksheet_xml(body)], [("Jumbled", "visible")])
    zip_parts("structure/out-of-order-rows.xlsx", parts)
    emit_sidecar("structure/out-of-order-rows.xlsx",
                 "Rows AND cells appear out of order, and <dimension> is missing entirely. "
                 "Any reader that assumes monotonic r or trusts dimension gets this wrong.",
                 [sheet("Jumbled",
                        {"A1": cell("number", 11.0), "B2": cell("number", 22.0),
                         "A7": cell("number", 71.0), "C7": cell("number", 73.0),
                         "B9": cell("number", 92.0), "D9": cell("number", 94.0),
                         "E9": cell("number", 95.0)},
                        dimension=None, used_range="A1:E9")],
                 verified_by=HAND)

    # 9. rich text runs in sharedStrings
    sst_xml = (XMLDECL + f'<sst xmlns="{NS_MAIN}" count="3" uniqueCount="3">'
               '<si><r><rPr><b/><sz val="11"/></rPr><t>Bold</t></r>'
               '<r><rPr><sz val="11"/></rPr><t xml:space="preserve"> then plain</t></r></si>'
               '<si><r><rPr><i/><color rgb="FFFF0000"/></rPr><t>red italic</t></r></si>'
               '<si><t xml:space="preserve">  leading and trailing  </t></si>'
               "</sst>")
    body = ('<row r="1"><c r="A1" t="s"><v>0</v></c><c r="B1" t="s"><v>1</v></c>'
            '<c r="C1" t="s"><v>2</v></c></row>')
    parts = base_parts([worksheet_xml(body, dimension="A1:C1")], [("Rich", "visible")])
    parts["xl/sharedStrings.xml"] = sst_xml
    zip_parts("structure/rich-text.xlsx", parts)
    emit_sidecar("structure/rich-text.xlsx",
                 "sharedStrings with multi-run rich text, and xml:space=\"preserve\" whitespace. "
                 "Runs flatten to plain text in the model; the raw part stays in OpaqueParts.",
                 [sheet("Rich",
                        {"A1": cell("text", "Bold then plain"),
                         "B1": cell("text", "red italic"),
                         "C1": cell("text", "  leading and trailing  ")},
                        dimension="A1:C1", used_range="A1:C1")],
                 verified_by=HAND,
                 passthrough=["xl/sharedStrings.xml"],
                 notes="C1 proves xml:space=preserve is honoured - trimming it is a silent "
                       "data change that survives into every later save.")

    # 10. hyperlinks
    sheet_rels = (XMLDECL + '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                  f'<Relationship Id="rId1" Type="{RT}/hyperlink" '
                  'Target="https://example.invalid/never-fetched" TargetMode="External"/>'
                  f'<Relationship Id="rId2" Type="{RT}/hyperlink" '
                  'Target="file:///etc/passwd" TargetMode="External"/>'
                  "</Relationships>")
    body = ('<row r="1"><c r="A1" t="inlineStr"><is><t>external link</t></is></c>'
            '<c r="B1" t="inlineStr"><is><t>file link</t></is></c>'
            '<c r="C1" t="inlineStr"><is><t>internal link</t></is></c>'
            '<c r="D1" t="str"><f>HYPERLINK("https://example.invalid/fn","fn link")</f>'
            "<v>fn link</v></c></row>")
    hl = ('<hyperlinks><hyperlink ref="A1" r:id="rId1"/><hyperlink ref="B1" r:id="rId2"/>'
          '<hyperlink ref="C1" location="Links!A100" display="jump"/></hyperlinks>')
    parts = base_parts([worksheet_xml(body, dimension="A1:D1", tail=hl)], [("Links", "visible")])
    parts["xl/worksheets/_rels/sheet1.xml.rels"] = sheet_rels
    zip_parts("structure/hyperlinks.xlsx", parts)
    emit_sidecar("structure/hyperlinks.xlsx",
                 "External http, file:// and internal-location hyperlinks plus a HYPERLINK() "
                 "formula. All four are inert until clicked and must show the resolved URL first.",
                 [sheet("Links",
                        {"A1": cell("text", "external link"),
                         "B1": cell("text", "file link"),
                         "C1": cell("text", "internal link"),
                         "D1": cell("text", "fn link",
                                    formula='HYPERLINK("https://example.invalid/fn","fn link")')},
                        dimension="A1:D1", used_range="A1:D1",
                        hyperlinks={"A1": "https://example.invalid/never-fetched",
                                    "B1": "file:///etc/passwd",
                                    "C1": "#Links!A100"})],
                 verified_by=HAND,
                 mustNotHappen=["opening file:///etc/passwd", "any network request"])

    # 11. the largest legal cell string
    big = ("x" * 32000) + ("ABCDEFGHIJ" * 76) + "END"
    assert len(big) == 32763, len(big)
    parts = base_parts([worksheet_xml(
        '<row r="1"><c r="A1" t="s"><v>0</v></c><c r="B1"><f>LEN(A1)</f><v>32763</v></c></row>',
        dimension="A1:B1")], [("Long", "visible")], sst=[big])
    zip_parts("structure/long-cell-32k.xlsx", parts)
    emit_sidecar("structure/long-cell-32k.xlsx",
                 "A 32,763-character cell - just under Excel's 32,767 limit, so it must be "
                 "ACCEPTED. Its hostile twin (40,000 chars) must be rejected.",
                 [sheet("Long",
                        {"A1": cell("text", None),
                         "B1": cell("number", 32763.0, formula="LEN(A1)")},
                        dimension="A1:B1", used_range="A1:B1")],
                 verified_by=HAND,
                 skip_checks=["cellValue:Long!A1"],
                 expectCellLength={"Long!A1": 32763},
                 notes="The sidecar asserts the LENGTH, not the 32 KB of text.")


# ==========================================================================
# passthrough/  --- the group that keeps A2 honest
# ==========================================================================

# Elements that live INSIDE xl/worksheets/sheetN.xml. They cannot be passed
# through byte-identically, because that is the very part the writer re-emits.
# A2 must either model them or splice the original fragments back in verbatim.
SHEET_LEVEL_SURVIVORS = [
    "conditionalFormatting", "dataValidations", "autoFilter", "sheetProtection",
    "pageMargins", "pageSetup", "printOptions", "headerFooter", "drawing",
    "legacyDrawing", "hyperlinks", "extLst", "tableParts", "phoneticPr",
]


def _xw(relpath, in_memory=True, **opts):
    import xlsxwriter
    opts.setdefault("in_memory", in_memory)
    return xlsxwriter.Workbook(fx(relpath), opts)


def _seed_grid(ws, rows=8):
    ws.write_row(0, 0, ["Region", "Q1", "Q2"])
    data = [("North", 120, 140), ("South", 95, 110), ("East", 78, 132),
            ("West", 210, 180), ("North", 60, 71), ("South", 44, 52),
            ("East", 155, 149)]
    for i, (r, a, b) in enumerate(data[:rows - 1], start=1):
        ws.write_row(i, 0, [r, a, b])
    return data[:rows - 1]


def g_passthrough():
    import xlsxwriter

    # 1. chart
    wb = _xw("passthrough/chart.xlsx"); ws = wb.add_worksheet("Data")
    _seed_grid(ws)
    ch = wb.add_chart({"type": "column"})
    ch.add_series({"name": "=Data!$B$1", "categories": "=Data!$A$2:$A$8",
                   "values": "=Data!$B$2:$B$8"})
    ch.add_series({"name": "=Data!$C$1", "values": "=Data!$C$2:$C$8"})
    ch.set_title({"name": "Quarterly"})
    ws.insert_chart("E2", ch)
    wb.close()
    emit_sidecar("passthrough/chart.xlsx",
                 "A real column chart. Its parts (chart1.xml, drawing1.xml and their rels) must "
                 "survive a cell edit byte-identical, and the <drawing r:id> element inside "
                 "sheet1.xml must be re-emitted or the chart vanishes from the sheet.",
                 [sheet("Data",
                        {"A1": cell("text", "Region"), "B1": cell("text", "Q1"),
                         "B2": cell("number", 120.0), "C8": cell("number", 149.0)},
                        used_range="A1:C8")],
                 verified_by=OPX,
                 sheetLevelElementsThatMustSurvive=["drawing", "pageMargins"],
                 staleRisk="Editing B2:C8 changes the chart's source values. The chart XML does "
                           "not cache them, so it stays correct - but a chart with a cached "
                           "<c:numCache> would go stale and must be warned about, not 'fixed'.")
    attach_passthrough_hashes("passthrough/chart.xlsx", [
        "xl/charts/chart1.xml", "xl/drawings/drawing1.xml",
        "xl/drawings/_rels/drawing1.xml.rels", "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/theme/theme1.xml", "docProps/app.xml", "docProps/core.xml"])

    # 2. image
    png = tiny_png()
    wb = _xw("passthrough/image.xlsx"); ws = wb.add_worksheet("Pictures")
    ws.write("A1", "an image lives at C3")
    ws.insert_image("C3", "swatch.png", {"image_data": io.BytesIO(png)})
    wb.close()
    emit_sidecar("passthrough/image.xlsx",
                 "An embedded PNG. xl/media/image1.png is binary and MUST be copied through as "
                 "already-deflated bytes - re-encoding it changes the file for no reason.",
                 [sheet("Pictures", {"A1": cell("text", "an image lives at C3")},
                        used_range="A1:A1")],
                 verified_by=OPX,
                 sheetLevelElementsThatMustSurvive=["drawing", "pageMargins"])
    attach_passthrough_hashes("passthrough/image.xlsx", [
        "xl/media/image1.png", "xl/drawings/drawing1.xml",
        "xl/drawings/_rels/drawing1.xml.rels", "xl/worksheets/_rels/sheet1.xml.rels"])

    # 3. conditional formatting  -- the trap
    wb = _xw("passthrough/conditional-format.xlsx"); ws = wb.add_worksheet("Rules")
    _seed_grid(ws)
    red = wb.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})
    ws.conditional_format("B2:C8", {"type": "cell", "criteria": ">", "value": 100, "format": red})
    ws.conditional_format("B2:C8", {"type": "3_color_scale"})
    ws.conditional_format("B2:B8", {"type": "data_bar", "data_bar_2010": True})
    ws.conditional_format("A2:A8", {"type": "duplicate", "format": red})
    wb.close()
    emit_sidecar("passthrough/conditional-format.xlsx",
                 "Four conditional-format rule kinds. THE POINT: <conditionalFormatting> lives "
                 "INSIDE sheet1.xml, the one part the writer re-emits, so it cannot be passed "
                 "through. A2 must model it or splice the original fragment back verbatim.",
                 [sheet("Rules",
                        {"A1": cell("text", "Region"), "B2": cell("number", 120.0)},
                        used_range="A1:C8")],
                 verified_by=OPX,
                 sheetLevelElementsThatMustSurvive=[
                     "conditionalFormatting", "extLst", "pageMargins"],
                 conditionalFormatRuleCount=4,
                 notes="The data-bar rule is written twice: a legacy <cfRule type=dataBar> AND a "
                       "modern <x14:conditionalFormatting> inside <extLst>. Dropping either one "
                       "loses the rule in some Excel versions. The dxf styles it points at live "
                       "in xl/styles.xml <dxfs>, which A2 also rewrites.")
    attach_passthrough_hashes("passthrough/conditional-format.xlsx", [
        "docProps/app.xml", "docProps/core.xml", "xl/theme/theme1.xml"])

    # 4. data validation  -- the same trap
    wb = _xw("passthrough/data-validation.xlsx"); ws = wb.add_worksheet("Validated")
    ws.write("A1", "pick one"); ws.write("B1", "1-100"); ws.write("C1", "a date")
    ws.data_validation("A2:A20", {"validate": "list",
                                  "source": ["alpha", "beta", "gamma"],
                                  "input_title": "Choose", "input_message": "One of three"})
    ws.data_validation("B2:B20", {"validate": "integer", "criteria": "between",
                                  "minimum": 1, "maximum": 100,
                                  "error_title": "Out of range",
                                  "error_message": "1 to 100 only"})
    ws.data_validation("C2:C20", {"validate": "date", "criteria": ">=",
                                  "value": "2024-01-01"})
    ws.data_validation("D2:D20", {"validate": "custom", "value": "=ISNUMBER(D2)"})
    wb.close()
    emit_sidecar("passthrough/data-validation.xlsx",
                 "List, integer-range, date and custom-formula validations. Like conditional "
                 "formatting these are IN-SHEET elements; losing them silently removes a "
                 "guardrail the file's author put there on purpose.",
                 [sheet("Validated",
                        {"A1": cell("text", "pick one"), "B1": cell("text", "1-100"),
                         "C1": cell("text", "a date")},
                        used_range="A1:C1")],
                 verified_by=OPX,
                 sheetLevelElementsThatMustSurvive=["dataValidations", "pageMargins"],
                 dataValidationCount=4)
    attach_passthrough_hashes("passthrough/data-validation.xlsx", [
        "docProps/app.xml", "docProps/core.xml"])

    # 5. comments (notes)
    wb = _xw("passthrough/comments.xlsx"); ws = wb.add_worksheet("Annotated")
    ws.write("A1", "hover me"); ws.write("B3", 42)
    ws.write_comment("A1", "This value came from the finance team.", {"author": "Fixture"})
    ws.write_comment("B3", "Do not edit.", {"author": "Fixture", "visible": True})
    wb.close()
    emit_sidecar("passthrough/comments.xlsx",
                 "Legacy comments. xl/comments1.xml and the VML that positions them are separate "
                 "parts (passthrough-able), but the <legacyDrawing r:id> pointer inside "
                 "sheet1.xml is not - drop it and the comments become orphans.",
                 [sheet("Annotated",
                        {"A1": cell("text", "hover me"), "B3": cell("number", 42.0)},
                        used_range="A1:B3")],
                 verified_by=OPX,
                 sheetLevelElementsThatMustSurvive=["legacyDrawing", "pageMargins"])
    attach_passthrough_hashes("passthrough/comments.xlsx", [
        "xl/comments1.xml", "xl/drawings/vmlDrawing1.vml",
        "xl/worksheets/_rels/sheet1.xml.rels"])

    # 6. macros
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        vba = os.path.join(td, "vbaProject.bin")
        with open(vba, "wb") as f:
            f.write(fake_vba_project())
        wb = _xw("passthrough/macros.xlsm")
        ws = wb.add_worksheet("Macro")
        ws.write("A1", "this workbook declares a VBA project")
        ws.write("A2", "OpenSheets never executes it")
        wb.add_vba_project(vba)
        wb.close()
    emit_sidecar("passthrough/macros.xlsm",
                 "xl/vbaProject.bin must survive byte-identical and must NEVER be executed "
                 "(PLAN.md 7.3). The app shows a 'contains macros, not executed' chip.",
                 [sheet("Macro",
                        {"A1": cell("text", "this workbook declares a VBA project"),
                         "A2": cell("text", "OpenSheets never executes it")},
                        used_range="A1:A2")],
                 verified_by=OPX,
                 containsMacros=True,
                 sheetLevelElementsThatMustSurvive=["pageMargins"],
                 notes="The .bin is a synthetic OLE2/CFB container, not compiled VBA. It is "
                       "inert test data: nothing in OpenSheets parses or runs it.")
    attach_passthrough_hashes("passthrough/macros.xlsm", ["xl/vbaProject.bin"])

    # 7. pivot table (hand-authored parts)
    cache_def = (XMLDECL + f'<pivotCacheDefinition xmlns="{NS_MAIN}" xmlns:r="{NS_R}" '
                 'r:id="rId1" refreshOnLoad="1" recordCount="4">'
                 '<cacheSource type="worksheet"><worksheetSource ref="A1:B5" sheet="Data"/></cacheSource>'
                 '<cacheFields count="2">'
                 '<cacheField name="Region" numFmtId="0"><sharedItems count="2">'
                 '<s v="North"/><s v="South"/></sharedItems></cacheField>'
                 '<cacheField name="Amount" numFmtId="0"><sharedItems containsSemiMixedTypes="0" '
                 'containsString="0" containsNumber="1" containsInteger="1" minValue="10" '
                 'maxValue="40"/></cacheField></cacheFields></pivotCacheDefinition>')
    cache_rec = (XMLDECL + f'<pivotCacheRecords xmlns="{NS_MAIN}" xmlns:r="{NS_R}" count="4">'
                 '<r><x v="0"/><n v="10"/></r><r><x v="1"/><n v="20"/></r>'
                 '<r><x v="0"/><n v="30"/></r><r><x v="1"/><n v="40"/></r>'
                 "</pivotCacheRecords>")
    pivot = (XMLDECL + f'<pivotTableDefinition xmlns="{NS_MAIN}" name="PivotTable1" '
             'cacheId="1" dataOnRows="1" applyNumberFormats="0" applyBorderFormats="0" '
             'applyFontFormats="0" applyPatternFormats="0" applyAlignmentFormats="0" '
             'applyWidthHeightFormats="1" dataCaption="Values" useAutoFormatting="1" '
             'itemPrintTitles="1" indent="0" outline="1" outlineData="1" multipleFieldFilters="0">'
             '<location ref="A3:B6" firstHeaderRow="1" firstDataRow="2" firstDataCol="0"/>'
             '<pivotFields count="2">'
             '<pivotField axis="axisRow" showAll="0"><items count="3">'
             '<item x="0"/><item x="1"/><item t="default"/></items></pivotField>'
             '<pivotField dataField="1" showAll="0"/></pivotFields>'
             '<rowFields count="1"><field x="0"/></rowFields>'
             '<rowItems count="3"><i><x/></i><i><x v="1"/></i>'
             '<i t="grand"><x/></i></rowItems>'
             '<colItems count="1"><i/></colItems>'
             '<dataFields count="1"><dataField name="Sum of Amount" fld="1" baseField="0" '
             'baseItem="0"/></dataFields></pivotTableDefinition>')
    pivot_rels = (XMLDECL + '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                  f'<Relationship Id="rId1" Type="{RT}/pivotCacheDefinition" '
                  'Target="../pivotCache/pivotCacheDefinition1.xml"/></Relationships>')
    cache_rels = (XMLDECL + '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                  f'<Relationship Id="rId1" Type="{RT}/pivotCacheRecords" '
                  'Target="pivotCacheRecords1.xml"/></Relationships>')
    sheet2_rels = (XMLDECL + '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                   f'<Relationship Id="rId1" Type="{RT}/pivotTable" '
                   'Target="../pivotTables/pivotTable1.xml"/></Relationships>')
    data_rows = ('<row r="1"><c r="A1" t="inlineStr"><is><t>Region</t></is></c>'
                 '<c r="B1" t="inlineStr"><is><t>Amount</t></is></c></row>'
                 '<row r="2"><c r="A2" t="inlineStr"><is><t>North</t></is></c><c r="B2"><v>10</v></c></row>'
                 '<row r="3"><c r="A3" t="inlineStr"><is><t>South</t></is></c><c r="B3"><v>20</v></c></row>'
                 '<row r="4"><c r="A4" t="inlineStr"><is><t>North</t></is></c><c r="B4"><v>30</v></c></row>'
                 '<row r="5"><c r="A5" t="inlineStr"><is><t>South</t></is></c><c r="B5"><v>40</v></c></row>')
    pivot_sheet = ('<row r="3"><c r="A3" t="inlineStr"><is><t>Row Labels</t></is></c>'
                   '<c r="B3" t="inlineStr"><is><t>Sum of Amount</t></is></c></row>'
                   '<row r="4"><c r="A4" t="inlineStr"><is><t>North</t></is></c><c r="B4"><v>40</v></c></row>'
                   '<row r="5"><c r="A5" t="inlineStr"><is><t>South</t></is></c><c r="B5"><v>60</v></c></row>'
                   '<row r="6"><c r="A6" t="inlineStr"><is><t>Grand Total</t></is></c><c r="B6"><v>100</v></c></row>')
    parts = base_parts(
        [worksheet_xml(data_rows, dimension="A1:B5"),
         worksheet_xml(pivot_sheet, dimension="A3:B6")],
        [("Data", "visible"), ("Pivot", "visible")])
    parts["xl/workbook.xml"] = workbook_xml(
        [("Data", "visible"), ("Pivot", "visible")],
        defined_names='<pivotCaches><pivotCache cacheId="1" r:id="rIdPc1"/></pivotCaches>')
    parts["xl/_rels/workbook.xml.rels"] = workbook_rels(
        2, extra=f'<Relationship Id="rIdPc1" Type="{RT}/pivotCacheDefinition" '
                 'Target="pivotCache/pivotCacheDefinition1.xml"/>')
    parts["[Content_Types].xml"] = content_types(2, extra=(
        '<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"/>'
        '<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"/>'
        '<Override PartName="/xl/pivotTables/pivotTable1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>'))
    parts["xl/pivotCache/pivotCacheDefinition1.xml"] = cache_def
    parts["xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels"] = cache_rels
    parts["xl/pivotCache/pivotCacheRecords1.xml"] = cache_rec
    parts["xl/pivotTables/pivotTable1.xml"] = pivot
    parts["xl/pivotTables/_rels/pivotTable1.xml.rels"] = pivot_rels
    parts["xl/worksheets/_rels/sheet2.xml.rels"] = sheet2_rels
    zip_parts("passthrough/pivot-table.xlsx", parts)
    emit_sidecar("passthrough/pivot-table.xlsx",
                 "A pivot table over Data!A1:B5 with its cache definition and cached records. "
                 "OpenSheets never authors pivots; it must render the flattened result on the "
                 "Pivot sheet and pass all four pivot parts through untouched.",
                 [sheet("Data",
                        {"A1": cell("text", "Region"), "B1": cell("text", "Amount"),
                         "B5": cell("number", 40.0)},
                        index=0, dimension="A1:B5", used_range="A1:B5"),
                  sheet("Pivot",
                        {"A3": cell("text", "Row Labels"), "B4": cell("number", 40.0),
                         "B5": cell("number", 60.0), "B6": cell("number", 100.0)},
                        index=1, dimension="A3:B6", used_range="A3:B6")],
                 verified_by=HAND,
                 sheetLevelElementsThatMustSurvive=[],
                 staleRisk="The pivot cache holds a COPY of Data!A1:B5. Editing Data does not "
                           "update it. refreshOnLoad=\"1\" is set so Excel rebuilds on open; "
                           "OpenSheets must warn rather than silently rewrite the cache.")
    attach_passthrough_hashes("passthrough/pivot-table.xlsx", [
        "xl/pivotCache/pivotCacheDefinition1.xml",
        "xl/pivotCache/pivotCacheRecords1.xml",
        "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
        "xl/pivotTables/pivotTable1.xml",
        "xl/pivotTables/_rels/pivotTable1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels"])

    # 8. everything at once
    with tempfile.TemporaryDirectory() as td:
        vba = os.path.join(td, "vbaProject.bin")
        with open(vba, "wb") as f:
            f.write(fake_vba_project())
        wb = _xw("passthrough/kitchen-sink.xlsm")
        ws = wb.add_worksheet("Everything")
        _seed_grid(ws)
        red = wb.add_format({"bg_color": "#FFC7CE"})
        ws.conditional_format("B2:C8", {"type": "cell", "criteria": ">", "value": 100,
                                        "format": red})
        ws.data_validation("A2:A8", {"validate": "list",
                                     "source": ["North", "South", "East", "West"]})
        ws.write_comment("A1", "region column", {"author": "Fixture"})
        ws.insert_image("H2", "swatch.png", {"image_data": io.BytesIO(png)})
        ch = wb.add_chart({"type": "line"})
        ch.add_series({"values": "=Everything!$B$2:$B$8"})
        ws.insert_chart("E2", ch)
        ws.add_table("A1:C8", {"name": "Sales", "columns": [
            {"header": "Region"}, {"header": "Q1"}, {"header": "Q2"}]})
        ws.write_url("A10", "https://example.invalid/never-fetched", None, "a link")
        ws.set_header("&LOpenSheets&RPage &P")
        ws.set_footer("&Cfixture")
        ws.print_row_col_headers()
        ws.fit_to_pages(1, 1)
        ws.set_paper(9)
        ws.freeze_panes(1, 1)
        ws.protect("", {"objects": False, "format_cells": True})
        wb.add_vba_project(vba)
        wb.close()
    emit_sidecar("passthrough/kitchen-sink.xlsm",
                 "Chart + image + conditional format + data validation + comment + autofilter + "
                 "sheet protection + frozen panes + VBA, all on one sheet. If A2 survives this "
                 "one intact it survives real workbooks.",
                 [sheet("Everything",
                        {"A1": cell("text", "Region"), "B2": cell("number", 120.0),
                         "C8": cell("number", 149.0),
                         "A10": cell("text", "a link")},
                        used_range="A1:C10",
                        frozen={"columns": 1, "rows": 1, "topLeftCell": "B2",
                                "state": "frozen"})],
                 verified_by=OPX,
                 containsMacros=True,
                 sheetLevelElementsThatMustSurvive=[
                     "conditionalFormatting", "dataValidations", "sheetProtection",
                     "hyperlinks", "printOptions", "pageMargins", "pageSetup",
                     "headerFooter", "drawing", "legacyDrawing", "tableParts"],
                 allSheetLevelElementsA2MustConsider=SHEET_LEVEL_SURVIVORS)
    attach_passthrough_hashes("passthrough/kitchen-sink.xlsm", [
        "xl/vbaProject.bin", "xl/media/image1.png", "xl/charts/chart1.xml",
        "xl/drawings/drawing1.xml", "xl/drawings/_rels/drawing1.xml.rels",
        "xl/drawings/vmlDrawing1.vml", "xl/comments1.xml", "xl/tables/table1.xml",
        "xl/worksheets/_rels/sheet1.xml.rels", "xl/theme/theme1.xml"])


# ==========================================================================
# csv/
# ==========================================================================

def g_csv():
    def put(name, data, proves, *, delimiter, quote, line_ending, encoding, bom,
            rows, ragged=False, **extra):
        write_bytes("csv/" + name, data)
        emit_csv_sidecar("csv/" + name, proves, delimiter=delimiter, quote=quote,
                         line_ending=line_ending, encoding=encoding, bom=bom,
                         rows=rows, ragged=ragged, **extra)

    rows = [["id", "name", "amount"], ["1", "Ada", "10.5"], ["2", "Grace", "-3"]]

    put("comma-lf.csv", b"id,name,amount\n1,Ada,10.5\n2,Grace,-3\n",
        "The baseline: comma, LF, trailing newline, no quoting needed.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False, rows=rows)

    put("semicolon-crlf.csv", b"id;name;amount\r\n1;Ada;10,5\r\n2;Grace;-3\r\n",
        "European dialect: semicolon delimiter, CRLF, comma as the DECIMAL separator. "
        "Sniffing the delimiter as ',' here splits 10,5 into two columns.",
        delimiter=";", quote='"', line_ending="\r\n", encoding="utf-8", bom=False,
        rows=[["id", "name", "amount"], ["1", "Ada", "10,5"], ["2", "Grace", "-3"]],
        decimalSeparator=",")

    put("tab.tsv", b"id\tname\tamount\n1\tAda\t10.5\n2\tGrace\t-3\n",
        "Tab-separated. RFC 4180 quoting still applies.",
        delimiter="\t", quote='"', line_ending="\n", encoding="utf-8", bom=False, rows=rows)

    put("pipe.csv", b"id|name|amount\n1|Ada|10.5\n2|Grace|-3\n",
        "Pipe delimiter - the fourth candidate the sniffer must consider.",
        delimiter="|", quote='"', line_ending="\n", encoding="utf-8", bom=False, rows=rows)

    put("cr-only.csv", b"id,name\r1,Ada\r2,Grace\r",
        "Classic Mac line endings (bare CR). Splitting on \\n yields ONE giant row.",
        delimiter=",", quote='"', line_ending="\r", encoding="utf-8", bom=False,
        rows=[["id", "name"], ["1", "Ada"], ["2", "Grace"]])

    put("quoted-newlines.csv",
        b'id,note\n1,"line one\nline two"\n2,"has,comma and\r\nCRLF inside"\n3,plain\n',
        "Newlines (LF and CRLF) inside quoted fields. A line-based reader corrupts this.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False,
        rows=[["id", "note"], ["1", "line one\nline two"],
              ["2", "has,comma and\r\nCRLF inside"], ["3", "plain"]])

    put("doubled-quotes.csv",
        b'id,text\n1,"she said ""hi"""\n2,"just ""one"" word"\n3,"""fully quoted"""\n',
        "RFC 4180 escaped quotes: \"\" inside a quoted field means one literal \".",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False,
        rows=[["id", "text"], ["1", 'she said "hi"'], ["2", 'just "one" word'],
              ["3", '"fully quoted"']])

    put("ragged-rows.csv", b"a,b,c\n1,2,3\n4,5\n6\n7,8,9,10,11\n",
        "Rows of 3, 3, 2, 1 and 5 fields. Short rows pad, long rows widen the sheet; "
        "neither is an error, but the count of ragged rows is reported.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False,
        rows=[["a", "b", "c"], ["1", "2", "3"], ["4", "5"], ["6"], ["7", "8", "9", "10", "11"]],
        ragged=True, raggedRowCount=3, maxColumnsSeen=5)

    put("bom-utf8.csv", b"\xef\xbb\xbfid,name\n1,\xc3\x84da\n",
        "UTF-8 BOM. It must be consumed, not delivered as part of the first field name.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=True,
        rows=[["id", "name"], ["1", "Äda"]])

    put("utf16le.csv", "id,name\n1,Ådä\n2,日本語\n".encode("utf-16-le")
        .rjust(len("id,name\n1,Ådä\n2,日本語\n".encode("utf-16-le")) + 2, b"\x00")
        .replace(b"\x00\x00", b"\xff\xfe", 1),
        "UTF-16LE with a BOM. Byte 2 of every ASCII char is NUL - a UTF-8 decoder either "
        "throws or produces interleaved NULs.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-16le", bom=True,
        rows=[["id", "name"], ["1", "Ådä"], ["2", "日本語"]])

    put("utf16be.csv", b"\xfe\xff" + "id,name\n1,Ådä\n".encode("utf-16-be"),
        "UTF-16BE with a BOM - the byte order the BOM must actually be read for.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-16be", bom=True,
        rows=[["id", "name"], ["1", "Ådä"]])

    put("windows-1252.csv",
        "id,name,note\n1,Ådä,naïve café\n2,Smart “quotes” and – dash,€99\n".encode("cp1252"),
        "Legacy Windows-1252 with no BOM and bytes that are INVALID UTF-8. The reader must "
        "fall back and tell the user it guessed.",
        delimiter=",", quote='"', line_ending="\n", encoding="windows-1252", bom=False,
        rows=[["id", "name", "note"], ["1", "Ådä", "naïve café"],
              ["2", "Smart “quotes” and – dash", "€99"]],
        mustSurfaceEncodingGuess=True)

    put("no-trailing-newline.csv", b"a,b\n1,2\n3,4",
        "The last line has no terminator. A reader that requires one drops the final row.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False,
        rows=[["a", "b"], ["1", "2"], ["3", "4"]])

    put("formula-injection.csv",
        b'label,payload\nequals,=1+1\nplus,+1+1\nminus,-1+1\nat,@SUM(A1)\n'
        b'cmd,"=cmd|\' /C calc\'!A0"\ntab,"\t=1+1"\ncr,"\r=1+1"\n'
        b'hyperlink,"=HYPERLINK(""http://x"",""click"")"\n',
        "Values that Excel would execute if this CSV were opened there. On EXPORT they must be "
        "prefixed with ' (PLAN.md 7.3); on IMPORT they stay inert text.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False,
        rows=[["label", "payload"], ["equals", "=1+1"], ["plus", "+1+1"],
              ["minus", "-1+1"], ["at", "@SUM(A1)"], ["cmd", "=cmd|' /C calc'!A0"],
              ["tab", "\t=1+1"], ["cr", "\r=1+1"],
              ["hyperlink", '=HYPERLINK("http://x","click")']],
        dangerousPrefixes=["=", "+", "-", "@", "\t", "\r"])

    put("empty.csv", b"",
        "Zero bytes. Must open as an empty sheet, not fail and not hang.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False, rows=[])

    put("header-only.csv", b"a,b,c\n",
        "One row, no data. usedRange is A1:C1.",
        delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False,
        rows=[["a", "b", "c"]])


# ==========================================================================
# perf/
# ==========================================================================

def g_perf(with_huge=False):
    import xlsxwriter

    def grid(relpath, rows, cols, proves, **extra):
        wb = xlsxwriter.Workbook(fx(relpath), {"constant_memory": True})
        ws = wb.add_worksheet("Perf")
        ws.write_row(0, 0, [f"col{c}" for c in range(cols)])
        for r in range(1, rows):
            ws.write_row(r, 0, [r * 1000 + c for c in range(cols)])
        wb.close()
        last = f"{get_column_letter(cols)}{rows}"
        emit_sidecar(relpath, proves,
                     [sheet("Perf",
                            {"A1": cell("text", "col0"),
                             "A2": cell("number", 1000.0),
                             last: cell("number", float((rows - 1) * 1000 + cols - 1))},
                            dimension=f"A1:{last}", used_range=f"A1:{last}")],
                     verified_by=OPX, **extra)

    grid("perf/100k-cells.xlsx", 1000, 100,
         "100,000 cells. Budget: open to first paint < 800 ms (PLAN.md 10.6).",
         cellCount=100000, budget={"openToFirstPaintMs": 800})

    grid("perf/wide-16384-cols.xlsx", 5, 16384,
         "All 16,384 columns x 5 rows. Column-width and style lookup must stay O(1) per cell "
         "and RunLengthArray must not allocate 16,384 doubles per sheet.",
         cellCount=81920, budget={"openMs": 2000})

    # a single cell at the very last legal address
    parts = base_parts([worksheet_xml(
        '<row r="1048576"><c r="XFD1048576"><v>1</v></c></row>',
        dimension="XFD1048576:XFD1048576")], [("Corner", "visible")])
    zip_parts("perf/single-cell-at-XFD1048576.xlsx", parts)
    emit_sidecar("perf/single-cell-at-XFD1048576.xlsx",
                 "One cell at the last legal address. usedRange is enormous but the sheet holds "
                 "ONE cell - anything that allocates per-cell over the used range dies here.",
                 [sheet("Corner", {"XFD1048576": cell("number", 1.0)},
                        dimension="XFD1048576:XFD1048576",
                        used_range="XFD1048576:XFD1048576")],
                 verified_by=HAND, cellCount=1,
                 budget={"openMs": 500, "peakRssMb": 100})

    if not with_huge:
        return

    grid("perf/1m-cells.xlsx", 10000, 100,
         "1,000,000 cells. Budget: open < 4 s, < 600 MB RSS (PLAN.md 10.6). Git-ignored: "
         "regenerate with Scripts/gen-fixtures.py perf --with-huge.",
         cellCount=1000000, budget={"openMs": 4000, "peakRssMb": 600}, generated=True)

    # a 2 GB CSV, streamed
    p = fx("perf/2gb.csv")
    chunk = "".join(f"{i},row-{i},{i*1.5},some padding text to widen the row\n"
                    for i in range(0, 2000))
    blob = chunk.encode()
    target = 2 * 1024 ** 3
    with open(p, "wb") as f:
        written = 0
        f.write(b"id,name,value,note\n")
        while written < target:
            f.write(blob)
            written += len(blob)
    emit_csv_sidecar("perf/2gb.csv",
                     "A 2 GB CSV. Must open with < 200 MB RSS - proof the reader streams "
                     "instead of slurping. Git-ignored; regenerate with --with-huge.",
                     delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False,
                     rows=[["id", "name", "value", "note"]],
                     generated=True, approxBytes=os.path.getsize(p),
                     budget={"peakRssMb": 200},
                     skipChecks=["rows"])

    # one line of 10 million characters
    p = fx("perf/single-line-10m.csv")
    with open(p, "wb") as f:
        f.write(b"header\n")
        f.write(b"x" * 10_000_000)
        f.write(b"\n")
    emit_csv_sidecar("perf/single-line-10m.csv",
                     "A single 10,000,000-character field. Bounds the per-field buffer; a "
                     "quadratic string append here takes minutes. Git-ignored.",
                     delimiter=",", quote='"', line_ending="\n", encoding="utf-8", bom=False,
                     rows=[["header"]],
                     generated=True, secondRowFieldLength=10_000_000,
                     skipChecks=["rows"])


# ==========================================================================
# hostile/   --- deliberately malformed. Never executed. Parser rejection tests.
# ==========================================================================

HOSTILE = []          # filled by _h(); written out as hostile/expected-errors.json


def _h(name, code, proves, *, also=(), must_not=(), notes=None, limit=None,
       expect_success=False):
    HOSTILE.append({
        "file": "hostile/" + name,
        "expectedError": None if expect_success else code,
        "expectedBehavior": code if expect_success else "throw",
        "alsoAcceptable": list(also),
        "mustNotHappen": list(must_not) or [
            "crash", "hang (>2s)", "unbounded memory growth",
            "reading any file outside the archive",
        ],
        "limit": limit,
        "proves": proves,
        "notes": notes,
    })


def _valid_base_bytes():
    import tempfile
    parts = base_parts([worksheet_xml('<row r="1"><c r="A1"><v>1</v></c></row>',
                                      dimension="A1:A1")], [("Sheet1", "visible")])
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for n, d in parts.items():
            zi = zipfile.ZipInfo(n, date_time=(2024, 1, 1, 0, 0, 0))
            zi.compress_type = zipfile.ZIP_DEFLATED
            z.writestr(zi, d.encode())
    return buf.getvalue(), parts


def _cfb(stream_names):
    """A minimal OLE2/CFB container naming the given streams."""
    blob = bytearray(fake_vba_project())
    # overwrite the directory entry names so the container advertises the
    # streams a password-protected .xlsx actually carries
    off = 512 * 2
    for i, nm in enumerate(stream_names):
        e = off + 128 * (i + 1)
        nb = nm.encode("utf-16-le") + b"\x00\x00"
        blob[e:e + 64] = nb.ljust(64, b"\x00")
        blob[e + 64:e + 66] = struct.pack("<H", len(nb))
        blob[e + 66] = 2
    return bytes(blob)


def g_hostile():
    base_bytes, base = _valid_base_bytes()

    # ---- 1. zip bomb on a part we MUST parse ----------------------------
    parts = dict(base)
    bomb = b"\x00" * (200 * 1024 * 1024)
    raw_zip("hostile/zip-bomb.xlsx",
            [{"name": n, "data": d.encode()} for n, d in parts.items()
             if n != "xl/worksheets/sheet1.xml"]
            + [{"name": "xl/worksheets/sheet1.xml", "data": bomb}])
    _h("zip-bomb.xlsx", "zip.compressionRatioExceeded",
       "xl/worksheets/sheet1.xml inflates ~1030:1 to 200 MB. The per-entry ratio cap must "
       "fire DURING inflation, not after.",
       also=["zip.decompressedSizeExceeded"],
       limit="Limits.maxCompressionRatio (100:1)",
       notes="Inflate incrementally and abort mid-stream; a reader that inflates to a buffer "
             "first has already lost 200 MB of RAM by the time it checks.")

    # ---- 2. bomb hidden in a part nobody should touch -------------------
    inner = io.BytesIO()
    with zipfile.ZipFile(inner, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("payload.bin", b"\x00" * (100 * 1024 * 1024))
    outer = io.BytesIO()
    with zipfile.ZipFile(outer, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("nested.zip", inner.getvalue())
    parts = dict(base)
    zip_parts("hostile/zip-bomb-nested.xlsx",
              {**parts, "xl/media/payload.zip": outer.getvalue()})
    _h("zip-bomb-nested.xlsx", "opensSuccessfullyWithoutInflatingUnmodelledParts",
       "A two-level zip bomb parked in xl/media/. The workbook is otherwise valid, so the "
       "CORRECT behaviour is to open it and never inflate that entry - OpaqueParts keeps the "
       "compressed bytes.",
       expect_success=True,
       must_not=["inflating xl/media/payload.zip", "recursing into nested archives",
                 "unbounded memory growth"],
       notes="This is the 'do not be eager' test. PLAN.md 5.2 says keep raw entries; A1's brief "
             "says 'do not decompress entries nobody asked for'. This file proves it.")

    # ---- 3. XXE ---------------------------------------------------------
    parts = dict(base)
    parts["xl/sharedStrings.xml"] = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<!DOCTYPE sst [\n'
        '  <!ENTITY xxe SYSTEM "file:///etc/passwd">\n'
        '  <!ENTITY xxe2 SYSTEM "http://127.0.0.1:1/ssrf">\n'
        "]>\n"
        f'<sst xmlns="{NS_MAIN}" count="2" uniqueCount="2">'
        "<si><t>&xxe;</t></si><si><t>&xxe2;</t></si></sst>")
    parts["xl/worksheets/sheet1.xml"] = worksheet_xml(
        '<row r="1"><c r="A1" t="s"><v>0</v></c><c r="B1" t="s"><v>1</v></c></row>',
        dimension="A1:B1")
    zip_parts("hostile/xxe-external-entity.xlsx", parts)
    _h("xxe-external-entity.xlsx", "xml.dtdForbidden",
       "sharedStrings declares external entities pointing at /etc/passwd and at a localhost "
       "URL. Neither may be resolved.",
       also=["xml.externalEntityForbidden"],
       must_not=["reading /etc/passwd", "any network connection",
                 "the string 'root:' appearing in any cell value", "crash", "hang (>2s)"],
       notes="XMLParser: shouldResolveExternalEntities = false AND externalEntityResolvingPolicy "
             "= .never. Setting only one of them is not enough.")

    # ---- 4. billion laughs ---------------------------------------------
    ents = ['<!ENTITY lol "lol">']
    for i in range(1, 10):
        prev = "lol" if i == 1 else f"lol{i-1}"
        ents.append(f'<!ENTITY lol{i} "&{prev};&{prev};&{prev};&{prev};&{prev};'
                    f'&{prev};&{prev};&{prev};&{prev};&{prev};">')
    parts = dict(base)
    parts["xl/sharedStrings.xml"] = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        "<!DOCTYPE sst [\n" + "\n".join(ents) + "\n]>\n"
        f'<sst xmlns="{NS_MAIN}" count="1" uniqueCount="1"><si><t>&lol9;</t></si></sst>')
    parts["xl/worksheets/sheet1.xml"] = worksheet_xml(
        '<row r="1"><c r="A1" t="s"><v>0</v></c></row>', dimension="A1:A1")
    zip_parts("hostile/billion-laughs.xlsx", parts)
    _h("billion-laughs.xlsx", "xml.dtdForbidden",
       "Nine levels of 10x entity expansion: &lol9; is 1,000,000,000 characters of 'lol'.",
       also=["xml.entityExpansionLimitExceeded"],
       limit="Limits.maxEntityExpansion",
       notes="If DTDs are refused outright this never gets a chance to expand, which is the "
             "outcome we want.")

    # ---- 5. a benign DOCTYPE -------------------------------------------
    parts = dict(base)
    parts["xl/sharedStrings.xml"] = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<!DOCTYPE sst SYSTEM "sst.dtd">\n'
        f'<sst xmlns="{NS_MAIN}" count="1" uniqueCount="1"><si><t>harmless</t></si></sst>')
    zip_parts("hostile/dtd-doctype.xlsx", parts)
    _h("dtd-doctype.xlsx", "xml.dtdForbidden",
       "A DOCTYPE with nothing malicious in it. The policy must be blanket - 'no DTDs in "
       "OOXML, ever' - not a heuristic that looks for the word SYSTEM.",
       notes="No legitimate xlsx producer emits a DOCTYPE. Rejecting all of them costs nothing.")

    # ---- 6/7. hostile entry names --------------------------------------
    entries = [{"name": n, "data": d.encode()} for n, d in base.items()]
    raw_zip("hostile/path-traversal.xlsx",
            entries + [{"name": "../../../../etc/passwd", "data": b"pwned\n"},
                       {"name": "xl/../../../tmp/evil.sh", "data": b"#!/bin/sh\n"},
                       {"name": "xl\\..\\..\\windows\\system32\\evil.dll", "data": b"MZ"}])
    _h("path-traversal.xlsx", "zip.entryNameTraversal",
       "Entry names containing ../ in POSIX and \\..\\ in Windows form. OpenSheets never "
       "extracts to disk, but OpaqueParts is keyed by entry name and A2 writes those names "
       "back out - a traversal name must never reach a file API.",
       limit="entry-name validation",
       notes="Reject on the name, before any inflation, and check the Windows separator too.")

    raw_zip("hostile/absolute-path-entry.xlsx",
            entries + [{"name": "/etc/passwd", "data": b"pwned\n"},
                       {"name": "C:\\Windows\\System32\\evil.dll", "data": b"MZ"}])
    _h("absolute-path-entry.xlsx", "zip.entryNameAbsolute",
       "Absolute POSIX and Windows entry names. Python's own zipfile.extract() strips these; "
       "a hand-rolled reader that does not is a write-anywhere primitive.")

    raw_zip("hostile/nul-in-entry-name.xlsx",
            entries + [{"name": "xl/worksheets/sheet1.xml\x00.png", "data": b"x"}])
    _h("nul-in-entry-name.xlsx", "zip.entryNameIllegalCharacter",
       "An entry name with an embedded NUL, so a C-string API sees "
       "'xl/worksheets/sheet1.xml' while the zip index sees something else.",
       notes="Classic poisoned-NUL confusion. Validate names as bytes, not as C strings.")

    # ---- 8. duplicate entry names --------------------------------------
    dup = [{"name": n, "data": d.encode()} for n, d in base.items()]
    dup.append({"name": "xl/worksheets/sheet1.xml",
                "data": worksheet_xml('<row r="1"><c r="A1"><v>666</v></c></row>',
                                      dimension="A1:A1").encode()})
    raw_zip("hostile/duplicate-entries.xlsx", dup)
    _h("duplicate-entries.xlsx", "zip.duplicateEntryName",
       "Two entries both named xl/worksheets/sheet1.xml with different content. Whichever one "
       "a reader picks, some other tool picks the other - so refuse the file.",
       notes="A1 must reject; A2 must never be able to EMIT a duplicate either.")

    # ---- 9/10. structurally broken archives ----------------------------
    write_bytes("hostile/truncated.xlsx", base_bytes[:int(len(base_bytes) * 0.55)])
    _h("truncated.xlsx", "zip.truncatedArchive",
       "The first 55% of a valid xlsx - no end-of-central-directory record at all. This is "
       "what a file caught mid-write looks like (PLAN.md 9), so it must be a clean, "
       "retryable error rather than a crash.",
       also=["zip.notAnArchive"])

    write_bytes("hostile/not-a-zip.xlsx",
                b"%PDF-1.7\n%\xe2\xe3\xcf\xd3\n1 0 obj\n<< /Type /Catalog >>\nendobj\n"
                + bytes(range(256)) * 8)
    _h("not-a-zip.xlsx", "zip.notAnArchive",
       "A PDF renamed to .xlsx. Trust the magic bytes, never the extension.")

    write_bytes("hostile/encrypted.xlsx",
                _cfb(["EncryptedPackage", "EncryptionInfo", "DataSpaces"]))
    _h("encrypted.xlsx", "xlsx.encryptedWorkbook",
       "An OLE2/CFB container advertising EncryptedPackage + EncryptionInfo - what Excel "
       "writes for a password-protected workbook. Must be distinguished from 'not a zip' so "
       "the UI can say 'password protected' instead of 'corrupt'.",
       also=["xlsx.unsupportedFormat"],
       notes="Detect on the CFB magic D0CF11E0A1B11AE1, then look for EncryptedPackage. "
             "Workbook.meta.readOnlyReason must be set so A2 can never write it back.")

    # ---- 11. entry-count bomb ------------------------------------------
    p = fx("hostile/entry-count-bomb.xlsx")
    with zipfile.ZipFile(p, "w", zipfile.ZIP_STORED) as z:
        for n, d in base.items():
            z.writestr(zipfile.ZipInfo(n, (2024, 1, 1, 0, 0, 0)), d.encode())
        for i in range(12000):
            z.writestr(zipfile.ZipInfo(f"xl/m/{i}", (2024, 1, 1, 0, 0, 0)), b"")
    _h("entry-count-bomb.xlsx", "zip.entryCountExceeded",
       "12,004 entries. Every one costs an OpaqueParts slot, so the cap has to be enforced "
       "while READING the central directory, not after building the table.",
       limit="Limits.maxZipEntries (10,000)")

    # ---- 12. an entry that lies about its size -------------------------
    raw_zip("hostile/lying-uncompressed-size.xlsx",
            [{"name": n, "data": d.encode()} for n, d in base.items()
             if n != "xl/sharedStrings.xml"]
            + [{"name": "xl/sharedStrings.xml",
                "data": shared_strings(["small"]).encode(),
                "declared_uncompressed": 10 * 1024 ** 3,
                "zip64": True}])
    _h("lying-uncompressed-size.xlsx", "zip.declaredSizeExceeded",
       "A ~120-byte entry whose ZIP64 header declares 10 GB uncompressed. Reject on the "
       "declared size BEFORE allocating, then verify the actual inflated length matches.",
       also=["zip.sizeMismatch"],
       limit="Limits.maxDecompressedBytes (500 MB)",
       notes="Both halves matter: pre-allocating from a declared size is a DoS, and trusting "
             "the declared size after inflating hides a truncation.")

    # ---- 13. CRC that does not match -----------------------------------
    raw_zip("hostile/crc-mismatch.xlsx",
            [{"name": n, "data": d.encode()} for n, d in base.items()
             if n != "xl/worksheets/sheet1.xml"]
            + [{"name": "xl/worksheets/sheet1.xml",
                "data": worksheet_xml('<row r="1"><c r="A1"><v>1</v></c></row>',
                                      dimension="A1:A1").encode(),
                "crc": 0xDEADBEEF}])
    _h("crc-mismatch.xlsx", "zip.checksumMismatch",
       "The stored CRC32 is wrong. Silent bit-rot on a modelled part must be caught, because "
       "A2 will copy the bad bytes straight back out on the next save.",
       notes="Verify the CRC of every entry you actually inflate. Entries passed through "
             "untouched keep their original CRC by construction.")

    # ---- 14. compression method we do not implement --------------------
    raw_zip("hostile/unsupported-compression-method.xlsx",
            [{"name": n, "data": d.encode()} for n, d in base.items()
             if n != "xl/worksheets/sheet1.xml"]
            + [{"name": "xl/worksheets/sheet1.xml", "data": b"whatever", "method": 99}])
    _h("unsupported-compression-method.xlsx", "zip.unsupportedCompressionMethod",
       "Method 99 = AES/WinZip encryption. Store (0) and deflate (8) are the only methods "
       "OOXML uses; anything else means the file is not what it claims.",
       notes="Report the method number in the message - 99 means encrypted, 12/14 mean "
             "bzip2/LZMA, and the user-facing text should differ.")

    # ---- 15. impossible dimensions -------------------------------------
    parts = dict(base)
    parts["xl/worksheets/sheet1.xml"] = worksheet_xml(
        '<row r="4294967295"><c r="A4294967295"><v>1</v></c></row>',
        dimension="A1:XFD4294967296")
    zip_parts("hostile/dimension-4-billion-rows.xlsx", parts)
    _h("dimension-4-billion-rows.xlsx", "xlsx.dimensionExceedsLimits",
       "<dimension> claims 4,294,967,296 rows and a row element sits at r=4294967295. "
       "Excel's ceiling is 1,048,576 x 16,384 (Limits).",
       also=["xlsx.cellReferenceOutOfRange"],
       limit="Limits.maxRows / Limits.maxColumns",
       notes="Also proves the row index is parsed as UInt32/Int64, not Int32 - 4294967295 "
             "overflows a signed 32-bit parse to -1.")

    # ---- 16. XML nested 100,000 deep ------------------------------------
    depth = 100000
    deep = (XMLDECL + f'<worksheet xmlns="{NS_MAIN}">'
            + "<a>" * depth + "</a>" * depth + "<sheetData/></worksheet>")
    parts = dict(base)
    parts["xl/worksheets/sheet1.xml"] = deep
    zip_parts("hostile/deep-nesting-100k.xlsx", parts)
    _h("deep-nesting-100k.xlsx", "xml.depthLimitExceeded",
       "100,000 levels of element nesting. A recursive-descent parser blows the stack; a "
       "DOM parser allocates 100,000 nodes. A depth cap stops both.",
       limit="Limits.maxXMLDepth",
       must_not=["stack overflow", "crash", "hang (>2s)", "unbounded memory growth"])

    # ---- 17. an over-length cell ---------------------------------------
    parts = dict(base)
    parts["xl/worksheets/sheet1.xml"] = worksheet_xml(
        '<row r="1"><c r="A1" t="inlineStr"><is><t>' + ("A" * 40000)
        + "</t></is></c></row>", dimension="A1:A1")
    zip_parts("hostile/cell-40k-chars.xlsx", parts)
    _h("cell-40k-chars.xlsx", "xlsx.cellTextTooLong",
       "A 40,000-character cell. Excel's hard limit is 32,767 (see the accepted twin "
       "structure/long-cell-32k.xlsx at 32,763).",
       limit="Limits.maxCellTextLength (32,767)",
       notes="If A1 would rather truncate-and-flag than reject, that is defensible - but it "
             "must be a deliberate, tested, user-visible choice, not an accident.")

    # ---- 18. NUL bytes inside a string ---------------------------------
    parts = dict(base)
    parts["xl/sharedStrings.xml"] = (
        XMLDECL + f'<sst xmlns="{NS_MAIN}" count="3" uniqueCount="3">'
        "<si><t>before\x00after</t></si>"
        "<si><t>bell\x07and\x0bvtab</t></si>"
        "<si><t>escaped _x0000_ form</t></si></sst>")
    parts["xl/worksheets/sheet1.xml"] = worksheet_xml(
        '<row r="1"><c r="A1" t="s"><v>0</v></c><c r="B1" t="s"><v>1</v></c>'
        '<c r="C1" t="s"><v>2</v></c></row>', dimension="A1:C1")
    zip_parts("hostile/nul-bytes-in-string.xlsx", parts)
    _h("nul-bytes-in-string.xlsx", "xml.illegalCharacter",
       "A raw NUL plus BEL and vertical tab inside <t>. All three are illegal in XML 1.0 and "
       "no conforming parser may accept them.",
       must_not=["crash", "hang (>2s)", "truncating the string at the NUL and continuing "
                 "silently"],
       notes="C1 carries the LEGAL escaped form _x0000_. If A1 decodes _x0000_ escapes it must "
             "strip the resulting control character rather than reintroducing it - and A2 must "
             "re-escape on write (its brief already calls this out as a real crash source).")

    # ---- 19. a required part is simply missing --------------------------
    parts = {k: v for k, v in base.items() if k != "xl/workbook.xml"}
    zip_parts("hostile/missing-workbook-part.xlsx", parts)
    _h("missing-workbook-part.xlsx", "xlsx.missingRequiredPart",
       "A structurally valid ZIP with [Content_Types].xml and rels but no xl/workbook.xml. "
       "The rels target dangles.",
       notes="Resolve parts THROUGH the rels graph (A1's brief: never hardcode paths) and this "
             "surfaces as a named missing relationship target, not an index-out-of-range.")

    # ---- 20. malformed cell references ---------------------------------
    parts = dict(base)
    parts["xl/worksheets/sheet1.xml"] = worksheet_xml(
        '<row r="0"><c r="A0"><v>1</v></c></row>'
        '<row r="1"><c r=""><v>2</v></c><c r="ZZZZZ1"><v>3</v></c>'
        '<c r="A"><v>4</v></c><c r="1A"><v>5</v></c>'
        '<c r="A-1"><v>6</v></c></row>', dimension="A1:B2")
    zip_parts("hostile/invalid-cell-reference.xlsx", parts)
    _h("invalid-cell-reference.xlsx", "xlsx.invalidCellReference",
       "Row 0 (rows are 1-based), an empty r=\"\", a 5-letter column past XFD, a column with "
       "no row, a row with no column, and a negative row.",
       also=["xlsx.cellReferenceOutOfRange"],
       notes="CellRef parsing is the most-called function in the codebase (A0's brief). It has "
             "to be both fast and total - every one of these must return nil, never trap.")

    # ---- 21. XML that is simply broken ---------------------------------
    parts = dict(base)
    parts["xl/worksheets/sheet1.xml"] = (
        XMLDECL + f'<worksheet xmlns="{NS_MAIN}"><sheetData><row r="1">'
        '<c r="A1"><v>1</v></row></sheetData>')
    zip_parts("hostile/malformed-xml.xlsx", parts)
    _h("malformed-xml.xlsx", "xml.malformed",
       "Unclosed <c>, unclosed <worksheet>, mismatched nesting. The baseline 'garbage in' case.",
       notes="The error must name the PART and ideally the byte offset. 'Could not open file' "
             "is not a diagnosis.")

    write_json(os.path.join(FIXTURES, "hostile", "expected-errors.json"), {
        "$comment": [
            "Ground truth for A1's parser-hardening suite (PLAN.md 7.4).",
            "Every file in Fixtures/hostile/ is deliberately malformed test data. None of it",
            "is ever executed. Run this suite under ASan with a 2 s per-case timeout.",
            "expectedError is the SheetError.code the reader MUST produce; alsoAcceptable",
            "lists codes that indicate the same defence firing at a different layer.",
            "One entry (zip-bomb-nested.xlsx) expects SUCCESS - read its notes.",
        ],
        "errorCodeNamespaces": {
            "zip.*": "MiniZip.Reader rejections - archive structure, entry names, size caps",
            "xml.*": "hardened XML pull-parser rejections - DTD, entities, depth, characters",
            "xlsx.*": "SheetFormat rejections - part graph, sheet limits, cell references",
        },
        "count": len(HOSTILE),
        "cases": HOSTILE,
    })


# ==========================================================================
# main
# ==========================================================================

GROUPS = {
    "basic": g_basic,
    "formulas": g_formulas,
    "formats": g_formats,
    "structure": g_structure,
    "passthrough": g_passthrough,
    "csv": g_csv,
    "hostile": g_hostile,
}


def main(argv=None):
    global LO
    ap = argparse.ArgumentParser(description="Generate the OpenSheets fixture corpus")
    ap.add_argument("groups", nargs="*", choices=sorted(GROUPS) + ["perf"], default=None)
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--with-huge", action="store_true",
                    help="also build the git-ignored 1M-cell xlsx, 2 GB csv and 10M-char line")
    args = ap.parse_args(argv)

    wanted = list(GROUPS) + ["perf"] if (args.all or not args.groups) else args.groups
    LO = soffice_version()

    for name in wanted:
        print(f"==> {name}", flush=True)
        if name == "perf":
            g_perf(with_huge=args.with_huge)
        else:
            GROUPS[name]()

    n_fx = sum(1 for root, _, files in os.walk(FIXTURES) for f in files
               if not f.endswith(".expected.json") and not f.endswith(".md")
               and f != "expected-errors.json")
    n_side = sum(1 for root, _, files in os.walk(FIXTURES) for f in files
                 if f.endswith(".expected.json"))
    print(f"\n{n_fx} fixtures, {n_side} sidecars")
    print(f"formula ground truth: {LO}")


if __name__ == "__main__":
    main()
